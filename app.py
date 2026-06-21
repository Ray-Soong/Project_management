from flask import Flask, jsonify, render_template, redirect, url_for, request, flash, send_file, session
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from models import CustomField, ProjectCustomFieldValue, db, Project, User, WorkLog, ProjectAssignment, ProjectManagerAssignment, Expense, ExpenseItem, Task, ProjectExpenseRecord, OperationLog, InvoiceFile
from forms import ProjectForm, LoginForm, UserForm, WorkLogForm, ProjectStatusForm, ExpenseForm, ExpenseApprovalForm, TaskForm, TaskUpdateForm, CustomFieldForm, UserEditForm
from datetime import datetime, date, timedelta
from utils import log_operation
import logging
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO

# 设置日志
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.config["SECRET_KEY"] = "yoursecret"
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///projects.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

# Session配置 - 10分钟未操作自动退出
app.config["PERMANENT_SESSION_LIFETIME"] = 600  # 600秒 = 10分钟
app.config["SESSION_REFRESH_EACH_REQUEST"] = False  # 不在每次请求时自动刷新，需要手动刷新

# 文件上传配置
UPLOAD_FOLDER = 'static/uploads'
CONTRACT_FOLDER = 'static/uploads/contracts'
INVOICE_FOLDER = 'static/uploads/invoices'
CONTRACT_ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'pdf', 'doc', 'docx', 'xls', 'xlsx', 'zip', 'rar', 'tar', 'gz', '7z'}
RECEIPT_ALLOWED_EXTENSIONS = {'zip'}
INVOICE_ALLOWED_EXTENSIONS = {'pdf', 'zip'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['CONTRACT_FOLDER'] = CONTRACT_FOLDER
app.config['INVOICE_FOLDER'] = INVOICE_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# 确保上传目录存在
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(CONTRACT_FOLDER, exist_ok=True)
os.makedirs(INVOICE_FOLDER, exist_ok=True)

def allowed_contract_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in CONTRACT_ALLOWED_EXTENSIONS

def allowed_invoice_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in INVOICE_ALLOWED_EXTENSIONS

def save_contract_file(file, project_id):
    """保存合同文件"""
    if file and allowed_contract_file(file.filename):
        # 生成安全的文件名
        filename = f"project_{project_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}"
        filepath = os.path.join(app.config['CONTRACT_FOLDER'], filename)
        file.save(filepath)
        return f"uploads/contracts/{filename}"
    return None

def save_invoice_file(file, project_id):
    """保存发票文件"""
    if file and allowed_invoice_file(file.filename):
        # 生成安全的文件名
        filename = f"invoice_{project_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}"
        filepath = os.path.join(app.config['INVOICE_FOLDER'], filename)
        file.save(filepath)
        return f"uploads/invoices/{filename}"
    return None

def save_invoice_files_to_db(project_id, files, user_id=None):
    """保存多个发票文件到数据库"""
    saved_files = []
    for file in files:
        if file and allowed_invoice_file(file.filename):
            try:
                # 生成安全的文件名
                filename = f"invoice_{project_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}"
                filepath = os.path.join(app.config['INVOICE_FOLDER'], filename)
                
                # 保存文件到磁盘
                file.save(filepath)
                
                # 获取文件大小
                file_size = os.path.getsize(filepath)
                
                # 获取文件扩展名
                file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else 'unknown'
                
                # 保存到数据库
                invoice_file = InvoiceFile(
                    project_id=project_id,
                    file_path=f"uploads/invoices/{filename}",
                    file_name=file.filename,
                    file_type=file_ext,
                    file_size=file_size,
                    uploaded_by=user_id
                )
                db.session.add(invoice_file)
                saved_files.append(invoice_file)
            except Exception as e:
                logger.error(f"保存发票文件失败: {e}")
                continue
    
    db.session.commit()
    return saved_files

# 登录管理器
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"

# 权限装饰器
from functools import wraps

def admin_or_finance_required(f):
    """需要管理员或财务权限"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('login'))
        if not current_user.has_full_access():
            flash("权限不足，只有管理员或财务可以访问")
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def admin_finance_or_pm_required(f):
    """需要管理员、财务或项目经理权限"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('login'))
        if not (current_user.has_full_access() or current_user.is_project_manager()):
            flash("权限不足，只有管理员、财务或项目经理可以访问")
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def check_project_access(f):
    """检查项目访问权限"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('login'))
        project_id = kwargs.get('project_id')
        if project_id and not current_user.can_view_project(project_id):
            flash("权限不足，您无法访问此项目")
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def create_expense_task(title, description, assigned_to, assigned_by, expense_id, priority='普通'):
    """创建报销处理任务，确保必须关联报销"""
    if not expense_id:
        raise ValueError("报销处理任务必须关联一个报销")
    
    # 验证expense是否存在
    expense = db.session.get(Expense, expense_id)
    if not expense:
        raise ValueError(f"报销ID不存在：{expense_id}")
    
    task = Task(
        title=title,
        description=description,
        task_type='expense_process',
        assigned_to=assigned_to,
        assigned_by=assigned_by,
        expense_id=expense_id,
        priority=priority
    )
    # 在添加到session前进行验证
    task.validate_expense_process_task()
    db.session.add(task)
    return task

login_manager.login_message = "请先登录"

db.init_app(app)

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

# 创建初始数据
def create_initial_data():
    # 创建管理员账户
    if not User.query.filter_by(username='admin').first():
        admin = User(username='admin', role='admin')
        admin.set_password('admin123')
        db.session.add(admin)
    
    # 创建开发者账户
    if not User.query.filter_by(username='developer').first():
        developer = User(username='developer', role='developer')
        developer.set_password('dev123')
        db.session.add(developer)
    
    db.session.commit()

# 检查用户活动并自动退出
@app.before_request
def check_session_timeout():
    """在每次请求前检查用户会话是否超时"""
    if current_user.is_authenticated:
        # 设置session为永久（这样才会使用PERMANENT_SESSION_LIFETIME配置）
        session.permanent = True
        
        # 检查上次活动时间
        last_activity = session.get('last_activity')
        now = datetime.now()
        
        if last_activity:
            last_activity_time = datetime.fromisoformat(last_activity)
            # 如果超过10分钟未活动，强制退出
            if now - last_activity_time > timedelta(seconds=600):
                # 记录超时退出
                username = current_user.username
                logout_user()
                session.clear()
                flash('由于长时间未操作，您已被自动退出，请重新登录', 'warning')
                return redirect(url_for('login'))
        
        # 更新最后活动时间（仅在非静态资源请求时更新）
        if not request.path.startswith('/static/'):
            session['last_activity'] = now.isoformat()

with app.app_context():
    # 注意：在生产环境中，应该使用数据库迁移而不是 drop_all()
    # 如果数据库不存在，创建它；如果存在，保留数据
    try:
        # 尝试查询用户表来检查数据库是否存在
        User.query.first()
        print("数据库已存在，保留现有数据")
        try:
            CustomField.query.first()
            print("自定义字段表已存在")
        except:
            print("创建自定义字段表...")
            # 只创建新表
            db.create_all()
            print("自定义字段表创建完成")
    except:
        # 如果查询失败，说明数据库不存在或表结构不对，重新创建
        print("创建新的数据库...")
        db.drop_all()
        db.create_all()
        create_initial_data()
        print("数据库初始化完成")

@app.route("/")
@login_required
def index():
    # 管理员和财务：查看所有项目
    # 项目经理：查看关联的项目
    # 开发工程师：跳转到工时列表
    
    if current_user.is_developer():
        return redirect(url_for('worklog_list'))
    
    # 获取筛选参数
    project_type = request.args.get('project_type')
    customer_name = request.args.get('customer_name')
    contract_date_from = request.args.get('contract_date_from')
    contract_date_to = request.args.get('contract_date_to')
    has_unpaid = request.args.get('has_unpaid')
    has_no_contract = request.args.get('has_no_contract')
    
    # 构建查询
    db.session.expire_all()
    query = Project.query
    
    # 根据用户角色限制可见项目
    if current_user.is_project_manager():
        # 项目经理只能看到关联的项目
        accessible_ids = current_user.get_accessible_project_ids()
        if not accessible_ids:
            # 如果没有关联项目，返回空列表
            projects = []
            stats = {
                'total_projects': 0,
                'active_projects': 0,
                'status_startup': 0,
                'status_paused': 0,
                'status_acceptance': 0,
                'status_acceptance_payment': 0,
                'total_contract_amount': 0,
                'total_cost': 0,
                'total_unpaid': 0,
                'gross_profit': 0
            }
            return render_template(
                "project_list.html",
                projects=projects,
                stats=stats,
                today=date.today(),
                project_types=[],
                customers=[],
                selected_type=project_type,
                selected_customer=customer_name,
                contract_date_from=contract_date_from,
                contract_date_to=contract_date_to,
                has_unpaid=has_unpaid,
                has_no_contract=has_no_contract
            )
        query = query.filter(Project.id.in_(accessible_ids))
    
    # 应用项目类别筛选
    if project_type:
        query = query.filter_by(project_type=project_type)
    
    # 应用客户筛选
    if customer_name:
        query = query.filter_by(customer_name=customer_name)
    
    # 应用合同日期范围筛选
    if contract_date_from:
        try:
            from_date = datetime.strptime(contract_date_from, '%Y-%m-%d').date()
            query = query.filter(Project.contract_signing_date >= from_date)
        except ValueError:
            pass
    
    if contract_date_to:
        try:
            to_date = datetime.strptime(contract_date_to, '%Y-%m-%d').date()
            query = query.filter(Project.contract_signing_date <= to_date)
        except ValueError:
            pass
    
    # 获取筛选后的项目列表
    projects = query.order_by(Project.project_number.asc().nullslast()).all()
    
    # 应用未收款筛选
    if has_unpaid == 'true':
        projects = [
            p for p in projects 
            if (p.contract_amount_with_tax or 0) - (p.payment_received or 0) > 0
        ]
    
    # 应用未签合同筛选
    if has_no_contract == 'true':
        projects = [
            p for p in projects 
            if p.contract_signing_date is None
        ]

    # 计算统计数据
    total_contract_amount = sum(
        float(p.contract_amount_with_tax) if p.contract_amount_with_tax else 0 
        for p in projects
    )
    
    total_cost = sum(p.get_total_cost() for p in projects)
    total_payment_received = sum(
        float(p.payment_received) if p.payment_received else 0
        for p in projects
    )
    # 未收款总和(含税) = 合同金额总和 - 已收款总和
    total_unpaid = total_contract_amount - total_payment_received
    # 应收款项总和
    total_accounts_receivable = sum(
        float(p.accounts_receivable) if p.accounts_receivable else 0
        for p in projects
    )
    gross_profit = total_payment_received / 1.06 - total_cost  # 毛利 = 回款金额/1.06 - 总成本
    active_projects_count = len([p for p in projects if p.status == '进行中'])
    
    # 统计各状态的项目数
    status_startup_count = len([p for p in projects if p.status == '启动中'])
    status_paused_count = len([p for p in projects if p.status == '暂停'])
    status_acceptance_count = len([p for p in projects if p.status == '验收中'])
    status_acceptance_payment_count = len([p for p in projects if p.status == '验收待回款'])
    
    stats = {
        'total_projects': len(projects),
        'active_projects': active_projects_count,
        'status_startup': status_startup_count,
        'status_paused': status_paused_count,
        'status_acceptance': status_acceptance_count,
        'status_acceptance_payment': status_acceptance_payment_count,
        'total_contract_amount': total_contract_amount,
        'total_cost': total_cost,
        'total_unpaid': total_unpaid,
        'gross_profit': gross_profit,
        'total_accounts_receivable': total_accounts_receivable
    }

    # 获取所有项目类别和客户用于筛选
    all_types = db.session.query(Project.project_type).distinct().filter(
        Project.project_type.isnot(None)
    ).all()
    project_types = [c[0] for c in all_types if c[0]]
    
    all_customers = db.session.query(Project.customer_name).distinct().filter(
        Project.customer_name.isnot(None)
    ).all()
    customers = [c[0] for c in all_customers if c[0]]
    
    # 获取所有活动的自定义字段
    custom_fields = CustomField.query.filter_by(is_active=True).all()

    today = date.today()
    
    return render_template(
        "project_list.html", 
        projects=projects, 
        stats=stats, 
        today=today,
        project_types=project_types,
        customers=customers,
        selected_type=project_type,
        selected_customer=customer_name,
        contract_date_from=contract_date_from,
        contract_date_to=contract_date_to,
        has_unpaid=has_unpaid,
        has_no_contract=has_no_contract,
        custom_fields=custom_fields
    )

@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("index"))
    
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user and user.check_password(form.password.data):
            login_user(user)
            # 设置session为永久并记录登录时间
            session.permanent = True
            session['last_activity'] = datetime.now().isoformat()
            # 记录登录日志
            log_operation('登录', '系统', f'用户 {user.username} 登录系统')
            next_page = request.args.get('next')
            return redirect(next_page) if next_page else redirect(url_for('index'))
        flash("用户名或密码错误")
    return render_template("login.html", form=form)

@app.route("/logout")
@login_required
def logout():
    # 记录登出日志
    log_operation('登出', '系统', f'用户 {current_user.username} 登出系统')
    logout_user()
    session.clear()
    return redirect(url_for("login"))

@app.route("/refresh_session", methods=["POST"])
@login_required
def refresh_session():
    """刷新session活动时间"""
    session['last_activity'] = datetime.now().isoformat()
    return jsonify({'status': 'success', 'message': 'Session已刷新'})

@app.route("/projects/export")
@login_required
def export_projects():
    """导出筛选后的项目列表为Excel文件"""
    # 权限检查：只有管理员、财务和项目经理可以导出
    if not (current_user.is_admin() or current_user.is_finance() or current_user.is_project_manager()):
        flash("权限不足，只有管理员、财务或项目经理可以导出项目")
        return redirect(url_for("index"))
    
    # 获取筛选参数（与index路由相同）
    project_type = request.args.get('project_type')
    customer_name = request.args.get('customer_name')
    contract_date_from = request.args.get('contract_date_from')
    contract_date_to = request.args.get('contract_date_to')  # 新增：结束日期
    has_unpaid = request.args.get('has_unpaid')
    has_no_contract = request.args.get('has_no_contract')
    
    # 构建查询（与index路由相同）
    query = Project.query
    
    # 根据用户角色限制可见项目
    if current_user.is_project_manager():
        # 项目经理只能导出关联的项目
        accessible_ids = current_user.get_accessible_project_ids()
        if not accessible_ids:
            # 如果没有关联项目，返回空的Excel文件
            projects = []
        else:
            query = query.filter(Project.id.in_(accessible_ids))
    
    if project_type:
        query = query.filter_by(project_type=project_type)
    
    if customer_name:
        query = query.filter_by(customer_name=customer_name)
    
    if contract_date_from:
        try:
            from_date = datetime.strptime(contract_date_from, '%Y-%m-%d').date()
            query = query.filter(Project.contract_signing_date >= from_date)
        except ValueError:
            pass
    
    if contract_date_to:
        try:
            to_date = datetime.strptime(contract_date_to, '%Y-%m-%d').date()
            query = query.filter(Project.contract_signing_date <= to_date)
        except ValueError:
            pass
        
    if current_user.is_project_manager() and not accessible_ids:
        projects = []
    else:
        projects = query.order_by(Project.project_number.asc().nullslast()).all()
    
    # 应用未收款筛选（需要在Python中过滤，因为这是计算字段）
    if has_unpaid == 'true':
        projects = [
            p for p in projects 
            if (p.contract_amount_with_tax or 0) - (p.payment_received or 0) > 0
        ]
    
    # 应用未签合同筛选
    if has_no_contract == 'true':
        projects = [
            p for p in projects 
            if p.contract_signing_date is None
        ]
    
    # 获取所有激活的自定义字段
    custom_fields = CustomField.query.filter_by(is_active=True).order_by(CustomField.id).all()
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "项目列表"
    
    # 定义样式
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border_side = Side(style='thin', color='000000')
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # 定义基础列标题
    headers = [
        "项目编号", "项目名称", "项目经理", "项目类型", "客户名称", "最终客户",
        "项目状态", "开始日期", "计划结束日期", "合同签订日期",
        "验收日期", "结算日期", "预计工时", "实际工时", "进度(%)",
        "合同金额(含税)", "合同金额(不含税)", "付款方式", "回款金额(含税)",
        "剩余金额", "发票状态", "开票日期", "开发费用", "报销费用",
        "外包费用(含税)", "供应商名称", "供应商发票是否开具", "外包费用备注", "间接成本", "间接成本备注", 
        "总成本", "毛利", "分配开发者", "创建时间"
    ]
    
    # 添加自定义字段到表头
    for custom_field in custom_fields:
        headers.append(custom_field.field_label)
    
    # 写入表头
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # 写入数据
    for row_num, project in enumerate(projects, 2):
        # 计算相关数据
        total_logged_hours = project.get_total_logged_hours()
        progress = project.get_progress_percentage()
        dev_cost = project.get_total_development_cost()
        
        # 报销费用
        expense_cost = 0
        project_expenses = Expense.query.filter_by(project_id=project.id, status='已批准').all()
        expense_cost = sum(float(e.total_amount) for e in project_expenses)
        
        # 外包费用(不含税)和间接成本
        outsourcing = float(project.outsourcing_cost_without_tax) if project.outsourcing_cost_without_tax else 0
        indirect = float(project.indirect_cost) if project.indirect_cost else 0
        
        # 总成本
        total_cost = dev_cost + expense_cost + outsourcing + indirect
        
        # 毛利 = 回款金额/1.06 - 总成本
        payment_received = float(project.payment_received) if project.payment_received else 0
        gross_profit = payment_received / 1.06 - total_cost
        
        # 合同金额(不含税)
        contract_without_tax = float(project.contract_amount_without_tax) if project.contract_amount_without_tax else 0
        
        # 分配的开发者
        assigned_devs = ", ".join([dev.username for dev in project.get_assigned_developers()])
        
        # 发票状态
        invoice_status = project.invoice_stage or "未开"
        
        # 剩余金额
        remaining = project.calculate_remaining_amount()
        
        # 数据行
        row_data = [
            project.id,
            project.name,
            project.manager,
            project.project_type or "",
            project.customer_name or "",
            project.final_customer or "",
            project.status,
            project.start_date.strftime('%Y-%m-%d') if project.start_date else "",
            project.planned_end_date.strftime('%Y-%m-%d') if project.planned_end_date else "",
            project.contract_signing_date.strftime('%Y-%m-%d') if project.contract_signing_date else "",
            project.acceptance_date.strftime('%Y-%m-%d') if project.acceptance_date else "",
            project.settlement_date.strftime('%Y-%m-%d') if project.settlement_date else "",
            project.estimated_hours or 0,
            total_logged_hours,
            f"{progress:.1f}",
            contract_with_tax,
            float(project.contract_amount_without_tax) if project.contract_amount_without_tax else 0,
            project.payment_method or "",
            float(project.payment_received) if project.payment_received else 0,
            remaining,
            invoice_status,
            project.invoice_date.strftime('%Y-%m-%d') if project.invoice_date else "",
            dev_cost,
            expense_cost,
            outsourcing,
            project.supplier_name or "",
            "是" if project.supplier_invoice_issued else "否",
            project.outsourcing_cost_notes or "",
            indirect,
            project.indirect_cost_notes or "",
            total_cost,
            gross_profit,
            assigned_devs,
            project.created_at.strftime('%Y-%m-%d %H:%M:%S') if project.created_at else ""
        ]
        
        # 添加自定义字段值
        for custom_field in custom_fields:
            field_value = ProjectCustomFieldValue.query.filter_by(
                project_id=project.id,
                custom_field_id=custom_field.id
            ).first()
            row_data.append(field_value.value if field_value else "")
        
        # 写入数据
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = cell_alignment
            cell.border = border
    
    # 调整列宽
    column_widths = {
        'A': 10, 'B': 25, 'C': 12, 'D': 15, 'E': 20,
        'F': 12, 'G': 12, 'H': 12, 'I': 12, 'J': 12,
        'K': 12, 'L': 10, 'M': 10, 'N': 10, 'O': 15,
        'P': 15, 'Q': 15, 'R': 12, 'S': 12, 'T': 12,
        'U': 12, 'V': 12, 'W': 12, 'X': 12, 'Y': 12,
        'Z': 25, 'AA': 12, 'AB': 12, 'AC': 20, 'AD': 20
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 生成文件名
    filename = f"项目列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # 记录操作日志
    log_operation('导出', '项目', f'导出了 {len(projects)} 个项目的详细信息')
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route("/projects/create", methods=["GET", "POST"])
@login_required
@admin_finance_or_pm_required
def create_project():
    form = ProjectForm()
    # 获取所有开发者用于选择
    developers = User.query.filter_by(role=User.ROLE_DEVELOPER).all()
    form.assigned_developers.choices = [(dev.id, dev.username) for dev in developers]
    
    # 获取所有项目经理用于选择
    project_managers = User.query.filter_by(role=User.ROLE_PROJECT_MANAGER).all()
    form.assigned_managers.choices = [(pm.id, pm.username) for pm in project_managers]
    
    if form.validate_on_submit():
        # 生成项目编号
        project_number = Project.generate_project_number()
        
        # 获取项目经理姓名（从分配的项目经理中获取）
        manager_names = []
        for manager_id in form.assigned_managers.data:
            manager = db.session.get(User, manager_id)
            if manager:
                manager_names.append(manager.username)
        manager_display = ", ".join(manager_names) if manager_names else "未分配"
        
        project = Project(
            project_number=project_number,
            name=form.name.data,
            manager=manager_display,  # 使用分配的项目经理姓名
            customer_name=form.customer_name.data,
            final_customer=form.final_customer.data,
            project_type=form.project_type.data,
            start_date=form.start_date.data,
            planned_end_date=form.planned_end_date.data,
            acceptance_date=form.acceptance_date.data,
            contract_signing_date=form.contract_signing_date.data,
            settlement_date=form.settlement_date.data,
            invoice_date=form.invoice_date.data,
            invoice_notes=form.invoice_notes.data,
            estimated_hours=form.estimated_hours.data,
            contract_amount_with_tax=form.contract_amount_with_tax.data,
            contract_amount_without_tax=form.contract_amount_without_tax.data,
            payment_method=form.payment_method.data,
            payment_received=form.payment_received.data or 0,
            invoice_stage=form.invoice_stage.data or '未开',
            status=form.status.data,
            outsourcing_cost=form.outsourcing_cost.data or 0,
            supplier_name=form.supplier_name.data,
            supplier_pending_amount=form.supplier_pending_amount.data or 0,
            outsourcing_cost_notes=form.outsourcing_cost_notes.data,
            indirect_cost=form.indirect_cost.data or 0,
            indirect_cost_notes=form.indirect_cost_notes.data,
            stage_payment_notes=form.stage_payment_notes.data,
            payment_amount_notes=form.payment_amount_notes.data,
            invoice_amount_issued=form.invoice_amount_issued.data or 0,
            current_invoice_amount=form.current_invoice_amount.data or 0,
            accounts_receivable=form.accounts_receivable.data or 0,
            po_number=form.po_number.data,
            area=form.area.data,
            unit_price=form.unit_price.data
        )
        # 计算剩余金额
        if project.contract_amount_with_tax and project.payment_received:
            project.remaining_amount = project.contract_amount_with_tax - project.payment_received
        elif project.contract_amount_with_tax:
            project.remaining_amount = project.contract_amount_with_tax
        
        db.session.add(project)
        db.session.flush()  # 获取项目ID
        
        # 处理合同文件上传
        if form.contract_file.data:
            contract_file = save_contract_file(form.contract_file.data, project.id)
            if contract_file:
                project.contract_file = contract_file
        
        # 分配开发者并记录名字
        assigned_devs = []
        for dev_id in form.assigned_developers.data:
            dev = db.session.get(User, dev_id)
            assignment = ProjectAssignment(project_id=project.id, user_id=dev_id)
            # 如果开发者有默认工时费用，自动设置
            if dev and dev.default_hourly_rate:
                assignment.hourly_rate = dev.default_hourly_rate
            db.session.add(assignment)
            assigned_devs.append(dev.username)
        
        # 分配项目经理并记录名字
        from models import ProjectManagerAssignment
        assigned_managers = []
        for manager_id in form.assigned_managers.data:
            assignment = ProjectManagerAssignment(project_id=project.id, user_id=manager_id)
            db.session.add(assignment)
            manager = db.session.get(User, manager_id)
            assigned_managers.append(manager.username)
        
        db.session.commit()

         # 记录操作日志
        log_operation(
            '创建',
            '项目',
            f'创建项目 "{project.name}"，客户：{project.customer_name or "未设置"}，分配工程师：{", ".join(assigned_devs) if assigned_devs else "无"}，项目经理：{", ".join(assigned_managers) if assigned_managers else "无"}',
            'project',
            project.id
        )

        flash("项目创建成功，请为开发者设置工时费用")
        return redirect(url_for("project_detail", project_id=project.id))
    return render_template("project_create.html", form=form)

@app.route("/projects/<int:project_id>")
@login_required
@check_project_access
def project_detail(project_id):
    """项目详情页面"""
    project = Project.query.get_or_404(project_id)
    developer_costs = project.get_developer_costs()
    total_dev_cost = project.get_total_development_cost()

    # 获取自定义字段
    custom_fields = CustomField.query.filter_by(is_active=True).all()
    
    # 获取当前项目的自定义字段值
    custom_field_values = {}
    for field in custom_fields:
        value = ProjectCustomFieldValue.query.filter_by(
            project_id=project_id, 
            custom_field_id=field.id
        ).first()
        if value:
            custom_field_values[field.id] = value.value

    # 获取该项目关联的所有报销单
    project_expenses = Expense.query.filter_by(project_id=project_id).order_by(Expense.submit_date.desc()).all()
    
    # 计算报销总金额
    total_expense_amount = sum(float(expense.total_amount) for expense in project_expenses if expense.status == '已批准')
    pending_expense_amount = sum(float(expense.total_amount) for expense in project_expenses if expense.status == '待审批')
    
    # 获取外包费用和间接成本
    outsourcing_cost = float(project.outsourcing_cost) if project.outsourcing_cost else 0
    indirect_cost = float(project.indirect_cost) if project.indirect_cost else 0
    
    # 计算总费用（开发费用 + 已批准的报销费用 + 外包费用(含税) + 间接成本）
    total_cost = total_dev_cost + total_expense_amount + outsourcing_cost + indirect_cost
    
    # 报销统计
    expense_stats = {
        'total_count': len(project_expenses),
        'approved_count': len([e for e in project_expenses if e.status == '已批准']),
        'pending_count': len([e for e in project_expenses if e.status == '待审批']),
        'rejected_count': len([e for e in project_expenses if e.status == '已拒绝']),
        'total_approved_amount': total_expense_amount,
        'total_pending_amount': pending_expense_amount
    }
            
    return render_template("project_detail.html", 
                         project=project, 
                         developer_costs=developer_costs,
                         total_dev_cost=total_dev_cost,
                         total_cost=total_cost,
                         custom_fields=custom_fields,
                         custom_field_values=custom_field_values,
                         project_expenses=project_expenses,
                         expense_stats=expense_stats)

@app.route('/custom-fields/manage', methods=['GET', 'POST'])
def manage_custom_fields():
    form = CustomFieldForm()
    
    if form.validate_on_submit():
        custom_field = CustomField(
            field_name=form.field_name.data,
            field_label=form.field_label.data,
            field_type=form.field_type.data,
            options=form.options.data if form.field_type.data == 'select' else None,
            is_required=form.is_required.data
        )
        db.session.add(custom_field)
        db.session.commit()
        flash('自定义字段添加成功！', 'success')
        return redirect(url_for('manage_custom_fields'))
    
    custom_fields = CustomField.query.all()
    return render_template('custom_fields_manage.html', form=form, custom_fields=custom_fields)

@app.route("/projects/<int:project_id>/edit", methods=["GET", "POST"])
@login_required
@admin_finance_or_pm_required
def edit_project(project_id):
    project = Project.query.get_or_404(project_id)

    # 记录修改前的所有信息用于对比
    old_data = {
        'name': project.name,
        'manager': project.manager,
        'customer_name': project.customer_name,
        'final_customer': project.final_customer,
        'project_type': project.project_type,
        'status': project.status,
        'estimated_hours': project.estimated_hours,
        'start_date': project.start_date,
        'planned_end_date': project.planned_end_date,
        'acceptance_date': project.acceptance_date,
        'contract_amount_with_tax': project.contract_amount_with_tax,
        'contract_amount_without_tax': project.contract_amount_without_tax,
        'payment_method': project.payment_method,
        'payment_received': project.payment_received,
        'contract_signing_date': project.contract_signing_date,
        'settlement_date': project.settlement_date,
        'invoice_date': project.invoice_date,
        'invoice_stage': project.invoice_stage,
        'invoice_amount': project.invoice_amount,
        'invoice_notes': project.invoice_notes,
        'outsourcing_cost': project.outsourcing_cost,
        'outsourcing_cost_without_tax': project.outsourcing_cost_without_tax,
        'supplier_name': project.supplier_name,
        'supplier_pending_amount': project.supplier_pending_amount,
        'outsourcing_cost_notes': project.outsourcing_cost_notes,
        'indirect_cost': project.indirect_cost,
        'indirect_cost_notes': project.indirect_cost_notes,
        'stage_payment_notes': project.stage_payment_notes,
        'payment_amount_notes': project.payment_amount_notes,
        'invoice_amount_issued': project.invoice_amount_issued,
        'current_invoice_amount': project.current_invoice_amount,
        'accounts_receivable': project.accounts_receivable,
        'po_number': project.po_number,
        'area': project.area,
        'unit_price': project.unit_price,
    }

    form = ProjectForm(obj=project)
    
    # 获取所有开发者用于选择
    developers = User.query.filter_by(role=User.ROLE_DEVELOPER).all()
    form.assigned_developers.choices = [(dev.id, dev.username) for dev in developers]
    
    # 获取所有项目经理用于选择
    project_managers = User.query.filter_by(role=User.ROLE_PROJECT_MANAGER).all()
    form.assigned_managers.choices = [(pm.id, pm.username) for pm in project_managers]
    
    # 获取自定义字段
    custom_fields = CustomField.query.filter_by(is_active=True).all()
    
    # 获取当前项目的自定义字段值
    custom_field_values = {}
    for field in custom_fields:
        value = ProjectCustomFieldValue.query.filter_by(
            project_id=project_id, 
            custom_field_id=field.id
        ).first()
        if value:
            custom_field_values[field.id] = value.value

    # 设置已分配的开发者和项目经理
    from models import ProjectManagerAssignment
    if request.method == 'GET':
        form.assigned_developers.data = [assignment.user_id for assignment in project.assignments]
        form.assigned_managers.data = [assignment.user_id for assignment in project.manager_assignments]
    
    if form.validate_on_submit():
        # 记录变更内容
        changes = []
        
        # 定义字段的中文名称映射
        field_labels = {
            'name': '项目名称',
            'manager': '项目经理',
            'customer_name': '客户名称',
            'final_customer': '最终客户',
            'project_type': '项目类型',
            'status': '项目状态',
            'estimated_hours': '预计工时',
            'start_date': '开始日期',
            'planned_end_date': '预计结束日期',
            'acceptance_date': '验收日期',
            'contract_amount_with_tax': '含税合同金额',
            'contract_amount_without_tax': '不含税合同金额',
            'payment_method': '回款方式',
            'payment_received': '回款金额(含税)',
            'contract_signing_date': '合同签订日期',
            'settlement_date': '结算日期',
            'invoice_date': '当前发票的开票日期',
            'invoice_stage': '当前发票开具阶段',
            'invoice_amount': '发票金额(含税)',
            'invoice_notes': '开票备注',
            'outsourcing_cost': '外包费用(含税)',
            'outsourcing_cost_without_tax': '外包费用(不含税)',
            'supplier_name': '供应商名称',
            'supplier_pending_amount': '供应商待付金额(不含税)',
            'supplier_invoice_issued': '供应商发票是否开具',
            'outsourcing_cost_notes': '外包费用备注',
            'indirect_cost': '间接成本',
            'indirect_cost_notes': '间接成本备注',
            'stage_payment_notes': '阶段付款备注',
            'payment_amount_notes': '回款金额备注',
            'invoice_amount_issued': '已开发票金额',
            'current_invoice_amount': '当前发票金额',
            'accounts_receivable': '应收账款',
            'po_number': 'PO号',
            'area': '面积',
            'unit_price': '单价',
        }

        # 格式化值的显示
        def format_value(field_name, value):
            if value is None:
                return '未设置'
            if field_name in ['start_date', 'planned_end_date', 'acceptance_date', 
                            'contract_signing_date', 'settlement_date', 'invoice_date']:
                return value.strftime('%Y-%m-%d') if value else '未设置'
            if field_name in ['contract_amount_with_tax', 'contract_amount_without_tax', 
                            'payment_received', 'outsourcing_cost', 'outsourcing_cost_without_tax', 'supplier_pending_amount', 'indirect_cost', 'invoice_amount', 
                            'invoice_amount_issued', 'current_invoice_amount', 'accounts_receivable', 'area', 'unit_price']:
                return f'¥{float(value):.2f}' if value else '¥0.00'
            if field_name == 'estimated_hours':
                return f'{float(value):.1f}小时' if value else '未设置'
            return str(value) if value else '未设置'
        
        # 获取将要设置的项目经理姓名
        manager_names = []
        for manager_id in form.assigned_managers.data:
            manager = db.session.get(User, manager_id)
            if manager:
                manager_names.append(manager.username)
        new_manager_value = ", ".join(manager_names) if manager_names else "未分配"
        
        # 对比基本字段变更
        form_data = {
            'name': form.name.data,
            'manager': new_manager_value,
            'customer_name': form.customer_name.data,
            'final_customer': form.final_customer.data,
            'project_type': form.project_type.data,
            'status': form.status.data,
            'estimated_hours': form.estimated_hours.data,
            'start_date': form.start_date.data,
            'planned_end_date': form.planned_end_date.data,
            'acceptance_date': form.acceptance_date.data,
            'contract_amount_with_tax': form.contract_amount_with_tax.data,
            'contract_amount_without_tax': form.contract_amount_without_tax.data,
            'payment_method': form.payment_method.data,
            'payment_received': form.payment_received.data or 0,
            'contract_signing_date': form.contract_signing_date.data,
            'settlement_date': form.settlement_date.data,
            'invoice_date': form.invoice_date.data,
            'invoice_stage': form.invoice_stage.data,
            'invoice_amount': form.invoice_amount.data or 0,
            'outsourcing_cost': form.outsourcing_cost.data or 0,
            'outsourcing_cost_without_tax': form.outsourcing_cost_without_tax.data or 0,
            'supplier_name': form.supplier_name.data,
            'supplier_pending_amount': form.supplier_pending_amount.data or 0,
            'outsourcing_cost_notes': form.outsourcing_cost_notes.data,
            'indirect_cost': form.indirect_cost.data or 0,
            'indirect_cost_notes': form.indirect_cost_notes.data,
            'stage_payment_notes': form.stage_payment_notes.data,
            'payment_amount_notes': form.payment_amount_notes.data,
            'invoice_amount_issued': form.invoice_amount_issued.data or 0,
            'current_invoice_amount': form.current_invoice_amount.data or 0,
            'accounts_receivable': form.accounts_receivable.data or 0,
            'po_number': form.po_number.data,
        }
        
        # 比较每个字段
        for field_name, new_value in form_data.items():
            old_value = old_data[field_name]
            # 特殊处理数字字段的比较
            if field_name in ['contract_amount_with_tax', 'contract_amount_without_tax', 
                            'payment_received', 'outsourcing_cost', 'outsourcing_cost_without_tax', 'supplier_pending_amount', 'indirect_cost', 'invoice_amount', 
                            'invoice_amount_issued', 'current_invoice_amount', 'accounts_receivable', 'estimated_hours']:
                old_val = float(old_value) if old_value else 0
                new_val = float(new_value) if new_value else 0
                if abs(old_val - new_val) > 0.01:  # 避免浮点数比较问题
                    changes.append(
                        f'{field_labels[field_name]}: {format_value(field_name, old_value)} → {format_value(field_name, new_value)}'
                    )
            elif old_value != new_value:
                changes.append(
                    f'{field_labels[field_name]}: {format_value(field_name, old_value)} → {format_value(field_name, new_value)}'
                )
        
        # 更新项目信息
        project.name = form.name.data
        # manager字段从分配的项目经理中获取
        manager_names = []
        for manager_id in form.assigned_managers.data:
            manager = db.session.get(User, manager_id)
            if manager:
                manager_names.append(manager.username)
        project.manager = ", ".join(manager_names) if manager_names else "未分配"
        
        project.customer_name = form.customer_name.data
        project.final_customer = form.final_customer.data
        project.project_type = form.project_type.data
        project.start_date = form.start_date.data
        project.planned_end_date = form.planned_end_date.data
        project.acceptance_date = form.acceptance_date.data
        project.contract_signing_date = form.contract_signing_date.data
        project.settlement_date = form.settlement_date.data
        project.invoice_date = form.invoice_date.data
        project.invoice_amount = form.invoice_amount.data
        project.invoice_notes = form.invoice_notes.data
        project.estimated_hours = form.estimated_hours.data
        project.contract_amount_with_tax = form.contract_amount_with_tax.data
        project.contract_amount_without_tax = form.contract_amount_without_tax.data
        project.payment_method = form.payment_method.data
        project.payment_received = form.payment_received.data or 0
        project.invoice_stage = form.invoice_stage.data or '未开'
        project.status = form.status.data
        # 处理外包费用(含税)和间接成本
        project.outsourcing_cost = form.outsourcing_cost.data or 0
        project.outsourcing_cost_without_tax = form.outsourcing_cost_without_tax.data or 0
        project.supplier_name = form.supplier_name.data
        project.supplier_invoice_issued = form.supplier_invoice_issued.data
        project.supplier_pending_amount = form.supplier_pending_amount.data or 0
        project.outsourcing_cost_notes = form.outsourcing_cost_notes.data
        project.indirect_cost = form.indirect_cost.data or 0
        project.indirect_cost_notes = form.indirect_cost_notes.data
        project.stage_payment_notes = form.stage_payment_notes.data
        project.payment_amount_notes = form.payment_amount_notes.data
        project.invoice_amount_issued = form.invoice_amount_issued.data or 0
        project.current_invoice_amount = form.current_invoice_amount.data or 0
        project.accounts_receivable = form.accounts_receivable.data or 0
        project.po_number = form.po_number.data
        
        # 处理合同文件上传
        if form.contract_file.data:
            # 删除旧文件
            if project.contract_file:
                old_file_path = os.path.join('static', project.contract_file)
                if os.path.exists(old_file_path):
                    os.remove(old_file_path)
            # 保存新文件
            contract_file = save_contract_file(form.contract_file.data, project.id)
            if contract_file:
                project.contract_file = contract_file
                changes.append(f'合同文件: 已更新')
        
        # 处理发票文件上传
        if form.invoice_file.data:
            # 删除旧文件
            if project.invoice_file:
                old_file_path = os.path.join('static', project.invoice_file)
                if os.path.exists(old_file_path):
                    os.remove(old_file_path)
            # 保存新文件
            invoice_file = save_invoice_file(form.invoice_file.data, project.id)
            if invoice_file:
                project.invoice_file = invoice_file
                changes.append(f'发票文件: 已更新')
        
        # 重新计算剩余金额
        if project.contract_amount_with_tax and project.payment_received:
            new_remaining = project.contract_amount_with_tax - project.payment_received
            if abs(float(project.remaining_amount or 0) - float(new_remaining)) > 0.01:
                changes.append(f'剩余金额: ¥{float(project.remaining_amount or 0):.2f} → ¥{float(new_remaining):.2f}')
            project.remaining_amount = project.contract_amount_with_tax - project.payment_received
        elif project.contract_amount_with_tax:
            project.remaining_amount = project.contract_amount_with_tax
        else:
            project.remaining_amount = 0

        # 获取当前分配的开发者
        current_assignments = {a.user_id: a for a in project.assignments}
        new_dev_ids = set(form.assigned_developers.data)
        current_dev_ids = set(current_assignments.keys())
        
        # 删除不再分配的开发者
        to_remove = current_dev_ids - new_dev_ids
        removed_devs = []
        for dev_id in to_remove:
            dev = db.session.get(User, dev_id)
            removed_devs.append(dev.username)
            db.session.delete(current_assignments[dev_id])
        
        # 添加新分配的开发者
        to_add = new_dev_ids - current_dev_ids
        added_devs = []
        for dev_id in to_add:
            dev = db.session.get(User, dev_id)
            added_devs.append(dev.username)
            assignment = ProjectAssignment(project_id=project.id, user_id=dev_id)
            db.session.add(assignment)

        # 记录人员变更
        if added_devs:
            changes.append(f'新增工程师: {", ".join(added_devs)}')
        if removed_devs:
            changes.append(f'移除工程师: {", ".join(removed_devs)}')
        
        # 处理项目经理分配
        from models import ProjectManagerAssignment
        current_manager_assignments = {a.user_id: a for a in project.manager_assignments}
        new_manager_ids = set(form.assigned_managers.data)
        current_manager_ids = set(current_manager_assignments.keys())
        
        # 删除不再分配的项目经理
        to_remove_managers = current_manager_ids - new_manager_ids
        removed_managers = []
        for manager_id in to_remove_managers:
            manager = db.session.get(User, manager_id)
            removed_managers.append(manager.username)
            db.session.delete(current_manager_assignments[manager_id])
        
        # 添加新分配的项目经理
        to_add_managers = new_manager_ids - current_manager_ids
        added_managers = []
        for manager_id in to_add_managers:
            manager = db.session.get(User, manager_id)
            added_managers.append(manager.username)
            assignment = ProjectManagerAssignment(project_id=project.id, user_id=manager_id)
            db.session.add(assignment)
        
        # 记录项目经理变更
        if added_managers:
            changes.append(f'新增项目经理: {", ".join(added_managers)}')
        if removed_managers:
            changes.append(f'移除项目经理: {", ".join(removed_managers)}')
        
        # 处理自定义字段
        for field in custom_fields:
            field_name = f'custom_field_{field.id}'
            old_custom_value = custom_field_values.get(field.id, '')

            if field_name in request.form:
                field_value = request.form[field_name]
                if old_custom_value != field_value:
                    changes.append(
                        f'{field.field_label}: {old_custom_value or "未设置"} → {field_value or "未设置"}'
                    )

                # 查找或创建自定义字段值记录
                custom_value = ProjectCustomFieldValue.query.filter_by(
                    project_id=project_id,
                    custom_field_id=field.id
                ).first()
                    
                if custom_value:
                    custom_value.value = field_value
                else:
                    custom_value = ProjectCustomFieldValue(
                        project_id=project_id,
                        custom_field_id=field.id,
                        value=field_value
                    )
                    db.session.add(custom_value)
            elif field.field_type == 'checkbox':
                # 复选框未选中时不会在 request.form 中
                new_custom_value = '0'
                if old_custom_value != '0' and old_custom_value != '':
                    changes.append(f'{field.field_label}: 选中 → 未选中')

                custom_value = ProjectCustomFieldValue.query.filter_by(
                    project_id=project_id,
                    custom_field_id=field.id
                ).first()
                
                if custom_value:
                    custom_value.value = '0'
                else:
                    custom_value = ProjectCustomFieldValue(
                        project_id=project_id,
                        custom_field_id=field.id,
                        value='0'
                    )
                    db.session.add(custom_value)

        db.session.commit()

        # 记录操作日志 - 更详细的变更记录
        if changes:
            change_detail = f'编辑项目 "{old_data["name"]}"，共修改 {len(changes)} 项内容：\n'
            change_detail += '\n'.join([f'  • {change}' for change in changes])
        else:
            change_detail = f'编辑项目 "{old_data["name"]}"（未做实际修改）'
        
        log_operation(
            '编辑',
            '项目',
            change_detail,
            'project',
            project.id
        )

        flash(f"项目更新成功，共修改了 {len(changes)} 项内容")
        return redirect(url_for("project_detail", project_id=project.id))
    
    return render_template("project_edit.html", 
                         form=form, 
                         project=project,
                         custom_fields=custom_fields,
                         custom_field_values=custom_field_values)

@app.route("/projects/<int:project_id>/autosave", methods=["POST"])
@login_required
@admin_finance_or_pm_required
def autosave_project(project_id):
    """自动保存项目编辑数据"""
    
    project = Project.query.get_or_404(project_id)
    
    try:
        # 获取表单数据
        data = request.get_json()
        
        # 更新项目基本信息
        if 'name' in data:
            project.name = data['name']
        if 'manager' in data:
            project.manager = data['manager']
        if 'customer_name' in data:
            project.customer_name = data['customer_name']
        if 'project_type' in data:
            project.project_type = data['project_type']
        if 'status' in data:
            project.status = data['status']
        if 'estimated_hours' in data:
            project.estimated_hours = float(data['estimated_hours']) if data['estimated_hours'] else None
        if 'start_date' in data:
            project.start_date = datetime.strptime(data['start_date'], '%Y-%m-%d').date() if data['start_date'] else None
        if 'planned_end_date' in data:
            project.planned_end_date = datetime.strptime(data['planned_end_date'], '%Y-%m-%d').date() if data['planned_end_date'] else None
        if 'acceptance_date' in data:
            project.acceptance_date = datetime.strptime(data['acceptance_date'], '%Y-%m-%d').date() if data['acceptance_date'] else None
        if 'contract_signing_date' in data:
            project.contract_signing_date = datetime.strptime(data['contract_signing_date'], '%Y-%m-%d').date() if data['contract_signing_date'] else None
        if 'settlement_date' in data:
            project.settlement_date = datetime.strptime(data['settlement_date'], '%Y-%m-%d').date() if data['settlement_date'] else None
        if 'invoice_date' in data:
            project.invoice_date = datetime.strptime(data['invoice_date'], '%Y-%m-%d').date() if data['invoice_date'] else None
        if 'invoice_notes' in data:
            project.invoice_notes = data['invoice_notes']
        if 'contract_amount_with_tax' in data:
            project.contract_amount_with_tax = float(data['contract_amount_with_tax']) if data['contract_amount_with_tax'] else None
        if 'contract_amount_without_tax' in data:
            project.contract_amount_without_tax = float(data['contract_amount_without_tax']) if data['contract_amount_without_tax'] else None
        if 'payment_method' in data:
            project.payment_method = data['payment_method']
        if 'payment_received' in data:
            project.payment_received = float(data['payment_received']) if data['payment_received'] else 0
        if 'invoice_stage' in data:
            project.invoice_stage = data['invoice_stage']
        if 'outsourcing_cost' in data:
            project.outsourcing_cost = float(data['outsourcing_cost']) if data['outsourcing_cost'] else 0
        if 'supplier_name' in data:
            project.supplier_name = data['supplier_name']
        if 'outsourcing_cost_notes' in data:
            project.outsourcing_cost_notes = data['outsourcing_cost_notes']
        if 'indirect_cost' in data:
            project.indirect_cost = float(data['indirect_cost']) if data['indirect_cost'] else 0
        if 'indirect_cost_notes' in data:
            project.indirect_cost_notes = data['indirect_cost_notes']
        
        # 重新计算剩余金额
        if project.contract_amount_with_tax and project.payment_received:
            project.remaining_amount = project.contract_amount_with_tax - project.payment_received
        elif project.contract_amount_with_tax:
            project.remaining_amount = project.contract_amount_with_tax
        else:
            project.remaining_amount = 0
        
        # 处理自定义字段
        if 'custom_fields' in data:
            for field_id, field_value in data['custom_fields'].items():
                custom_value = ProjectCustomFieldValue.query.filter_by(
                    project_id=project_id,
                    custom_field_id=int(field_id)
                ).first()
                
                if custom_value:
                    custom_value.value = field_value
                else:
                    custom_value = ProjectCustomFieldValue(
                        project_id=project_id,
                        custom_field_id=int(field_id),
                        value=field_value
                    )
                    db.session.add(custom_value)
        
        # 提交数据库更改
        db.session.commit()
        
        # 刷新项目对象以确保获取最新数据
        db.session.refresh(project)
        
        return jsonify({
            'success': True,
            'message': '自动保存成功',
            'timestamp': datetime.now().strftime('%H:%M:%S'),
            'project_name': project.name,
            'project_status': project.status
        })
    
    except Exception as e:
        db.session.rollback()
        logger.error(f"自动保存失败: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'自动保存失败: {str(e)}'
        }), 500

@app.route("/projects/<int:project_id>/contract/download")
@login_required
def download_contract_file(project_id):
    """下载合同文件"""
    if not current_user.is_admin():
        flash("权限不足")
        return redirect(url_for("index"))
    
    project = Project.query.get_or_404(project_id)
    if not project.contract_file:
        flash("该项目没有上传合同文件")
        return redirect(url_for("project_detail", project_id=project_id))
    
    file_path = os.path.join('static', project.contract_file)
    if not os.path.exists(file_path):
        flash("文件不存在")
        return redirect(url_for("project_detail", project_id=project_id))
    
    return send_file(file_path, as_attachment=True)

@app.route("/projects/<int:project_id>/contract/delete", methods=["POST"])
@login_required
def delete_contract_file(project_id):
    """删除合同文件"""
    if not current_user.is_admin():
        return jsonify({"success": False, "message": "权限不足"}), 403
    
    project = Project.query.get_or_404(project_id)
    if not project.contract_file:
        return jsonify({"success": False, "message": "没有文件可删除"})
    
    # 删除文件
    file_path = os.path.join('static', project.contract_file)
    if os.path.exists(file_path):
        try:
            os.remove(file_path)
        except Exception as e:
            return jsonify({"success": False, "message": f"删除文件失败: {str(e)}"})
    
    # 清除数据库记录
    project.contract_file = None
    db.session.commit()
    
    # 记录操作日志
    log_operation('删除', '项目', f'删除项目 "{project.name}" 的合同文件', 'project', project.id)
    
    return jsonify({"success": True, "message": "文件删除成功"})

@app.route("/projects/<int:project_id>/invoice/download")
@login_required
def download_invoice_file(project_id):
    """下载发票文件"""
    if not current_user.is_admin() and not current_user.is_finance():
        flash("权限不足")
        return redirect(url_for("index"))
    
    project = Project.query.get_or_404(project_id)
    if not project.invoice_file:
        flash("该项目没有上传发票文件")
        return redirect(url_for("project_detail", project_id=project_id))
    
    file_path = os.path.join('static', project.invoice_file)
    if not os.path.exists(file_path):
        flash("文件不存在")
        return redirect(url_for("project_detail", project_id=project_id))
    
    return send_file(file_path, as_attachment=True)

@app.route("/projects/<int:project_id>/invoice/delete", methods=["POST"])
@login_required
def delete_invoice_file(project_id):
    """删除发票文件"""
    if not current_user.is_admin() and not current_user.is_finance():
        return jsonify({"success": False, "message": "权限不足"}), 403
    
    project = Project.query.get_or_404(project_id)
    if not project.invoice_file:
        return jsonify({"success": False, "message": "没有文件可删除"})
    
    # 删除文件
    file_path = os.path.join('static', project.invoice_file)
    if os.path.exists(file_path):
        try:
            os.remove(file_path)
        except Exception as e:
            return jsonify({"success": False, "message": f"删除文件失败: {str(e)}"})
    
    # 清除数据库记录
    project.invoice_file = None
    db.session.commit()
    
    # 记录操作日志
    log_operation('删除', '项目', f'删除项目 "{project.name}" 的发票文件', 'project', project.id)
    
    return jsonify({"success": True, "message": "文件删除成功"})

@app.route("/projects/<int:project_id>/invoice_files/upload", methods=["POST"])
@login_required
@admin_or_finance_required
def upload_invoice_files(project_id):
    """上传多个发票文件"""
    project = Project.query.get_or_404(project_id)
    
    if 'files' not in request.files:
        return jsonify({"success": False, "message": "没有选择文件"})
    
    files = request.files.getlist('files')
    if not files:
        return jsonify({"success": False, "message": "没有选择文件"})
    
    try:
        saved_files = save_invoice_files_to_db(project_id, files, current_user.id)
        if saved_files:
            log_operation('上传', '项目', f'上传 {len(saved_files)} 个发票文件到项目 "{project.name}"', 'project', project.id)
            return jsonify({
                "success": True, 
                "message": f"成功上传 {len(saved_files)} 个文件",
                "files": [{
                    "id": f.id,
                    "name": f.file_name,
                    "type": f.file_type,
                    "size": f.file_size,
                    "upload_date": f.upload_date.strftime('%Y-%m-%d %H:%M:%S')
                } for f in saved_files]
            })
        else:
            return jsonify({"success": False, "message": "没有有效的文件"})
    except Exception as e:
        logger.error(f"上传发票文件失败: {e}")
        return jsonify({"success": False, "message": f"上传失败: {str(e)}"})

@app.route("/invoice_files/<int:file_id>/download")
@login_required
def download_invoice_file_by_id(file_id):
    """下载指定的发票文件"""
    if not current_user.is_admin() and not current_user.is_finance():
        flash("权限不足")
        return redirect(url_for("index"))
    
    invoice_file = InvoiceFile.query.get_or_404(file_id)
    project = invoice_file.project
    
    file_path = os.path.join('static', invoice_file.file_path)
    if not os.path.exists(file_path):
        flash("文件不存在")
        return redirect(url_for("project_detail", project_id=project.id))
    
    return send_file(file_path, as_attachment=True, download_name=invoice_file.file_name)

@app.route("/invoice_files/<int:file_id>/delete", methods=["POST"])
@login_required
def delete_invoice_file_by_id(file_id):
    """删除指定的发票文件"""
    if not current_user.is_admin() and not current_user.is_finance():
        return jsonify({"success": False, "message": "权限不足"}), 403
    
    invoice_file = InvoiceFile.query.get_or_404(file_id)
    project = invoice_file.project
    
    # 删除文件
    file_path = os.path.join('static', invoice_file.file_path)
    if os.path.exists(file_path):
        try:
            os.remove(file_path)
        except Exception as e:
            return jsonify({"success": False, "message": f"删除文件失败: {str(e)}"})
    
    # 从数据库删除记录
    db.session.delete(invoice_file)
    db.session.commit()
    
    # 记录操作日志
    log_operation('删除', '项目', f'删除项目 "{project.name}" 的发票文件: {invoice_file.file_name}', 'project', project.id)
    
    return jsonify({"success": True, "message": "文件删除成功"})

@app.route("/projects/<int:project_id>/invoice_files", methods=["GET"])
@login_required
def get_invoice_files(project_id):
    """获取项目的所有发票文件列表"""
    if not current_user.is_admin() and not current_user.is_finance():
        return jsonify({"success": False, "message": "权限不足"}), 403
    
    project = Project.query.get_or_404(project_id)
    
    files = InvoiceFile.query.filter_by(project_id=project_id, is_deleted=False).all()
    
    return jsonify({
        "success": True,
        "files": [{
            "id": f.id,
            "name": f.file_name,
            "type": f.file_type,
            "size": f.file_size,
            "upload_date": f.upload_date.strftime('%Y-%m-%d %H:%M:%S'),
            "uploaded_by": f.uploader.username if f.uploader else "未知"
        } for f in files]
    })

@app.route("/projects/<int:project_id>/assignment/<int:assignment_id>/rate", methods=["POST"])
@login_required
def update_hourly_rate(project_id, assignment_id):
    """更新开发者工时费用"""
    if not current_user.is_admin():
        return jsonify({"error": "权限不足"}), 403
    
    assignment = ProjectAssignment.query.get_or_404(assignment_id)
    if assignment.project_id != project_id:
        return jsonify({"error": "数据不匹配"}), 400
    
    hourly_rate = request.json.get('hourly_rate')
    if hourly_rate is not None:
        try:
            assignment.hourly_rate = float(hourly_rate) if hourly_rate != '' else None
            db.session.commit()
            return jsonify({"success": True, "message": "工时费用更新成功"})
        except ValueError:
            return jsonify({"error": "无效的费用格式"}), 400
    
    return jsonify({"error": "缺少费用数据"}), 400

@app.route("/projects/<int:project_id>/status", methods=["GET", "POST"])
@login_required
@admin_finance_or_pm_required
def update_project_status(project_id):
    """快速更新项目状态"""
    
    project = Project.query.get_or_404(project_id)
    form = ProjectStatusForm()
    
    # 设置当前状态
    if request.method == 'GET':
        form.status.data = project.status
    
    if form.validate_on_submit():
        old_status = project.status
        project.status = form.status.data
        db.session.commit()
        flash(f"项目状态已从 '{old_status}' 更新为 '{project.status}'")
        return redirect(url_for("project_detail", project_id=project.id))
    
    return render_template("project_status_update.html", form=form, project=project)

@app.route("/users/create", methods=["GET", "POST"])
@login_required
@admin_or_finance_required
def create_user():
    form = UserForm()
    if form.validate_on_submit():
        if User.query.filter_by(username=form.username.data).first():
            flash("用户名已存在")
            return render_template("user_create.html", form=form)
        
        new_user = User(username=form.username.data, role=form.role.data)
        new_user.set_password(form.password.data)
        
        # 设置默认工时费用
        if form.default_hourly_rate.data:
            new_user.default_hourly_rate = form.default_hourly_rate.data
        
        db.session.add(new_user)
        db.session.commit()
        
        # 记录操作日志
        hourly_rate_info = f'，默认工时费用：{form.default_hourly_rate.data}元/小时' if form.default_hourly_rate.data else ''
        log_operation('创建', '用户', f'创建用户 "{new_user.username}"，角色：{new_user.role}{hourly_rate_info}')
        
        flash("用户创建成功")
        return redirect(url_for("user_list"))
    return render_template("user_create.html", form=form)

@app.route("/users")
@login_required
@admin_or_finance_required
def user_list():
    """用户列表"""
    users = User.query.order_by(User.created_at.desc()).all()
    return render_template("user_list.html", users=users)

@app.route("/users/<int:user_id>/edit", methods=["GET", "POST"])
@login_required
@admin_or_finance_required
def edit_user(user_id):
    """编辑用户信息"""
    user = User.query.get_or_404(user_id)
    form = UserEditForm()
    
    if request.method == "GET":
        # 预填充表单数据
        form.username.data = user.username
        form.default_hourly_rate.data = user.default_hourly_rate
    
    if form.validate_on_submit():
        # 检查用户名是否被其他用户占用
        existing_user = User.query.filter(User.username == form.username.data, User.id != user_id).first()
        if existing_user:
            flash("用户名已存在，请选择其他用户名")
            return render_template("user_edit.html", form=form, user=user)
        
        # 检查密码确认
        if form.password.data:
            if form.password.data != form.confirm_password.data:
                flash("两次输入的密码不一致")
                return render_template("user_edit.html", form=form, user=user)
            user.set_password(form.password.data)
            log_operation('修改', '用户', f'修改用户 "{user.username}" 的密码')
        
        # 更新用户信息
        old_username = user.username
        user.username = form.username.data
        user.default_hourly_rate = form.default_hourly_rate.data
        
        db.session.commit()
        
        # 记录操作日志
        if old_username != user.username:
            log_operation('修改', '用户', f'将用户名从 "{old_username}" 修改为 "{user.username}"')
        log_operation('修改', '用户', f'更新用户 "{user.username}" 的信息')
        
        flash(f"用户 {user.username} 的信息已更新")
        return redirect(url_for("user_list"))
    
    return render_template("user_edit.html", form=form, user=user)

@app.route("/users/<int:user_id>/reset_password", methods=["GET", "POST"])
@login_required
@admin_or_finance_required
def reset_user_password(user_id):
    """管理员重置用户密码（保留旧功能作为兼容）"""
    return redirect(url_for('edit_user', user_id=user_id))

@app.route("/worklogs")
@login_required
def worklog_list():
    """工时记录列表"""
    # 获取查询参数
    user_id = request.args.get('user_id', type=int)
    project_id = request.args.get('project_id', type=int)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # 基础查询
    query = WorkLog.query
    
    # 根据用户角色过滤
    if current_user.is_developer():
        # 开发工程师只能看自己的工时
        query = query.filter_by(user_id=current_user.id)
    elif current_user.is_project_manager():
        # 项目经理只能看关联项目的工时
        accessible_project_ids = current_user.get_accessible_project_ids()
        if accessible_project_ids:
            query = query.filter(WorkLog.project_id.in_(accessible_project_ids))
        else:
            # 如果没有关联项目，返回空查询
            query = query.filter(WorkLog.id == -1)
    
    # 应用过滤条件
    if user_id:
        query = query.filter_by(user_id=user_id)
    
    if project_id:
        query = query.filter_by(project_id=project_id)
    
    # 应用日期范围过滤
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            query = query.filter(WorkLog.date >= start_date_obj)
        except ValueError:
            pass
    
    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
            query = query.filter(WorkLog.date <= end_date_obj)
        except ValueError:
            pass
    
    # 获取结果并排序
    worklogs = query.order_by(WorkLog.date.desc(), WorkLog.created_at.desc()).all()
    
    # 获取所有用户和项目用于下拉选择
    # 管理员、财务和项目经理都可以看到所有开发人员
    users = User.query.filter_by(role=User.ROLE_DEVELOPER).all() if (current_user.has_full_access() or current_user.is_project_manager()) else []
    
    # 根据用户角色获取可访问的项目列表
    if current_user.has_full_access():
        projects = Project.query.all()
    elif current_user.is_project_manager():
        accessible_project_ids = current_user.get_accessible_project_ids()
        projects = Project.query.filter(Project.id.in_(accessible_project_ids)).all() if accessible_project_ids else []
    else:
        # 开发工程师看到自己分配的项目
        projects = Project.query.join(ProjectAssignment).filter(
            ProjectAssignment.user_id == current_user.id
        ).all()
    
    return render_template(
        "worklog_list.html", 
        worklogs=worklogs,
        users=users,
        projects=projects,
        selected_user_id=user_id,
        selected_project_id=project_id,
        start_date=start_date,
        end_date=end_date
    )

@app.route("/worklogs/export", methods=["GET"])
@login_required
def export_worklogs():
    """导出工时记录为Excel"""
    # 获取查询参数
    user_id = request.args.get('user_id', type=int)
    project_id = request.args.get('project_id', type=int)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # 基础查询 - 与worklog_list相同的权限检查
    query = WorkLog.query
    
    # 根据用户角色过滤
    if current_user.is_developer():
        # 开发工程师只能看自己的工时
        query = query.filter_by(user_id=current_user.id)
    elif current_user.is_project_manager():
        # 项目经理只能看关联项目的工时
        accessible_project_ids = current_user.get_accessible_project_ids()
        if accessible_project_ids:
            query = query.filter(WorkLog.project_id.in_(accessible_project_ids))
        else:
            query = query.filter(WorkLog.id == -1)
    
    # 应用过滤条件
    if user_id:
        query = query.filter_by(user_id=user_id)
    
    if project_id:
        query = query.filter_by(project_id=project_id)
    
    # 应用日期范围过滤
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            query = query.filter(WorkLog.date >= start_date_obj)
        except ValueError:
            pass
    
    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
            query = query.filter(WorkLog.date <= end_date_obj)
        except ValueError:
            pass
    
    # 获取结果并排序
    worklogs = query.order_by(WorkLog.date.desc(), WorkLog.created_at.desc()).all()
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "工时记录"
    
    # 定义样式
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 设置列宽
    column_widths = {
        'A': 10,  # ID
        'B': 15,  # 开发者
        'C': 20,  # 项目名称
        'D': 12,  # 工作日期
        'E': 10,  # 工作时长
        'F': 40,  # 工作描述
        'G': 15,  # 记录时间
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 写入表头
    headers = ["记录ID", "开发者", "项目名称", "工作日期", "工作时长(小时)", "工作描述", "记录时间"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # 写入数据
    total_hours = 0
    for row_idx, log in enumerate(worklogs, 2):
        ws.cell(row=row_idx, column=1).value = log.id
        ws.cell(row=row_idx, column=2).value = log.user.username
        ws.cell(row=row_idx, column=3).value = log.project.name
        ws.cell(row=row_idx, column=4).value = log.date.strftime('%Y-%m-%d')
        ws.cell(row=row_idx, column=5).value = log.hours
        ws.cell(row=row_idx, column=6).value = log.description or ""
        ws.cell(row=row_idx, column=7).value = log.created_at.strftime('%Y-%m-%d %H:%M:%S')
        
        total_hours += log.hours
        
        # 应用边框和对齐
        for col_idx in range(1, 8):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            if col_idx == 4:  # 日期列
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 5:  # 时长列
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = '0.0'
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # 添加总计行
    if worklogs:
        last_row = len(worklogs) + 2
        ws.cell(row=last_row, column=1).value = "总计"
        ws.cell(row=last_row, column=1).font = Font(bold=True, size=11)
        ws.cell(row=last_row, column=5).value = total_hours
        ws.cell(row=last_row, column=5).font = Font(bold=True, size=11)
        ws.cell(row=last_row, column=5).number_format = '0.0'
        
        # 总计行样式
        for col_idx in range(1, 8):
            cell = ws.cell(row=last_row, column=col_idx)
            cell.fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
            cell.border = border
            if col_idx == 5:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # 冻结表头
    ws.freeze_panes = "A2"
    
    # 保存到字节流
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # 生成文件名 - 包含导出时间和查询条件
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"工时记录_导出_{timestamp}.xlsx"
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route("/worklogs/create", methods=["GET", "POST"])
@login_required
def create_worklog():
    if not current_user.is_developer():
        flash("只有开发工程师可以记录工时")
        return redirect(url_for("index"))

    form = WorkLogForm()
    # 只显示分配给当前开发者的项目
    assigned_projects = db.session.query(Project).join(ProjectAssignment).filter(
        ProjectAssignment.user_id == current_user.id
    ).all()
    
    if not assigned_projects:
        flash("您还没有被分配到任何项目，请联系项目经理")
        return redirect(url_for("worklog_list"))
    
    form.project_id.choices = [(p.id, p.name) for p in assigned_projects]
    
    if form.validate_on_submit():
        worklog = WorkLog(
            user_id=current_user.id,
            project_id=form.project_id.data,
            date=form.date.data,
            hours=form.hours.data,
            description=form.description.data
        )
        db.session.add(worklog)
        db.session.commit()

        # 获取项目信息用于日志 - 添加这两行
        project = db.session.get(Project, form.project_id.data)
        log_operation(
            '创建',
            '工时',
            f'记录工时 {form.hours.data}小时，项目：{project.name}',
            'worklog',
            worklog.id
        )

        flash("工时记录成功")
        return redirect(url_for("worklog_list"))
    
    return render_template("worklog_create.html", form=form)

@app.route("/worklogs/<int:worklog_id>/edit", methods=["GET", "POST"])
@login_required
def edit_worklog(worklog_id):
    """编辑工时记录"""
    worklog = WorkLog.query.get_or_404(worklog_id)
    
    # 权限检查：只有记录创建者或管理员可以编辑
    if worklog.user_id != current_user.id and not current_user.is_admin():
        flash("您没有权限编辑此工时记录")
        return redirect(url_for("worklog_list"))
    
    # 记录修改前的数据
    old_data = {
        'project_id': worklog.project_id,
        'date': worklog.date,
        'hours': worklog.hours,
        'description': worklog.description
    }
    
    form = WorkLogForm(obj=worklog)
    
    # 设置项目选择列表
    if current_user.is_admin():
        # 管理员可以选择所有项目
        all_projects = Project.query.all()
        form.project_id.choices = [(p.id, p.name) for p in all_projects]
    else:
        # 普通用户只能选择分配给自己的项目
        assigned_projects = db.session.query(Project).join(ProjectAssignment).filter(
            ProjectAssignment.user_id == current_user.id
        ).all()
        form.project_id.choices = [(p.id, p.name) for p in assigned_projects]
    
    if form.validate_on_submit():
        # 记录变更
        changes = []
        if old_data['project_id'] != form.project_id.data:
            old_project = db.session.get(Project, old_data['project_id'])
            new_project = db.session.get(Project, form.project_id.data)
            changes.append(f'项目: {old_project.name} → {new_project.name}')
        
        if old_data['date'] != form.date.data:
            changes.append(f'日期: {old_data["date"].strftime("%Y-%m-%d")} → {form.date.data.strftime("%Y-%m-%d")}')
        
        if old_data['hours'] != form.hours.data:
            changes.append(f'工时: {old_data["hours"]}小时 → {form.hours.data}小时')
        
        if old_data['description'] != form.description.data:
            changes.append(f'描述已更新')
        
        # 更新工时记录
        worklog.project_id = form.project_id.data
        worklog.date = form.date.data
        worklog.hours = form.hours.data
        worklog.description = form.description.data
        
        db.session.commit()
        
        # 记录操作日志
        if changes:
            change_detail = f'编辑工时记录 #{worklog_id}，变更: {", ".join(changes)}'
        else:
            change_detail = f'编辑工时记录 #{worklog_id}（未做实际修改）'
        
        log_operation(
            '编辑',
            '工时',
            change_detail,
            'worklog',
            worklog.id
        )
        
        flash("工时记录更新成功")
        return redirect(url_for("worklog_list"))
    
    return render_template("worklog_edit.html", form=form, worklog=worklog)

def allowed_receipt_file(filename):
    """检查报销材料文件扩展名是否允许"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in RECEIPT_ALLOWED_EXTENSIONS

def build_approver_choices(project_id=None, include_placeholder=True):
    role_labels = {
        User.ROLE_ADMIN: '管理员',
        User.ROLE_FINANCE: '财务',
        User.ROLE_PROJECT_MANAGER: '项目经理'
    }
    choices = []
    if include_placeholder:
        choices.append(('', '请选择审批人'))

    seen_ids = set()
    admin_finance = User.query.filter(User.role.in_([User.ROLE_ADMIN, User.ROLE_FINANCE])).all()
    for user in admin_finance:
        if user.id in seen_ids:
            continue
        choices.append((user.id, f"{user.username} ({role_labels.get(user.role, user.role)})"))
        seen_ids.add(user.id)

    if project_id:
        pm_assignments = ProjectManagerAssignment.query.filter_by(
            project_id=project_id
        ).order_by(ProjectManagerAssignment.assigned_date.asc()).all()
        for assignment in pm_assignments:
            user = assignment.user
            if not user or user.id in seen_ids:
                continue
            choices.append((user.id, f"{user.username} ({role_labels.get(user.role, user.role)})"))
            seen_ids.add(user.id)

    return choices

def get_default_approver_id(expense):
    if expense.project_id:
        assignment = ProjectManagerAssignment.query.filter_by(
            project_id=expense.project_id
        ).order_by(ProjectManagerAssignment.assigned_date.asc()).first()
        if assignment:
            return assignment.user_id

    admin_user = User.query.filter_by(role=User.ROLE_ADMIN).order_by(User.id.asc()).first()
    if admin_user:
        return admin_user.id

    finance_user = User.query.filter_by(role=User.ROLE_FINANCE).order_by(User.id.asc()).first()
    if finance_user:
        return finance_user.id

    return None

# 报销相关路由
@app.route("/api/approvers")
@login_required
def approver_options():
    project_id = request.args.get('project_id', type=int)
    if project_id and not (current_user.has_full_access() or current_user.can_view_project(project_id)):
        return jsonify({'error': 'forbidden'}), 403

    choices = build_approver_choices(project_id, include_placeholder=False)
    return jsonify([{'id': item[0], 'label': item[1]} for item in choices])

@app.route("/expenses/backfill-tasks", methods=["POST"])
@login_required
def expense_backfill_tasks():
    if not current_user.has_full_access():
        flash("只有管理员和财务可以补建任务")
        return redirect(url_for("expense_list"))

    pending_expenses = Expense.query.filter_by(status='待审批').all()
    created_count = 0

    for expense in pending_expenses:
        existing_task = Task.query.filter_by(
            expense_id=expense.id,
            task_type='expense_process'
        ).first()
        if existing_task:
            continue

        approver_id = get_default_approver_id(expense)
        if not approver_id:
            continue

        task = Task(
            title=f"审批报销：{expense.title}",
            description=f"报销金额：¥{expense.total_amount}\n报销说明：{expense.description or '无'}",
            task_type='expense_process',
            assigned_to=approver_id,
            assigned_by=current_user.id,
            expense_id=expense.id,
            priority='普通'
        )
        # 验证任务
        task.validate_expense_process_task()
        db.session.add(task)
        created_count += 1

    db.session.commit()
    flash(f"已补建 {created_count} 条待审批任务")
    return redirect(url_for("expense_list"))

@app.route("/expenses")
@login_required
def expense_list():
    """报销列表"""
    # 获取筛选参数
    user_id = request.args.get('user_id', type=int)
    expense_type = request.args.get('expense_type')
    project_id = request.args.get('project_id', type=int)
    submit_date_from = request.args.get('submit_date_from')
    submit_date_to = request.args.get('submit_date_to')
    expense_category = request.args.get('expense_category')
    
    # 构建查询
    if current_user.has_full_access():
        # 管理员和财务可以看到所有报销
        query = Expense.query
    elif current_user.is_project_manager():
        # 项目经理可以看到关联项目的报销
        accessible_project_ids = current_user.get_accessible_project_ids()
        if accessible_project_ids:
            # 包含关联项目的报销和没有关联项目的售前费用
            query = Expense.query.filter(
                db.or_(
                    Expense.project_id.in_(accessible_project_ids),
                    Expense.project_id.is_(None)
                )
            )
        else:
            query = Expense.query.filter(Expense.project_id.is_(None))
    else:
        # 开发工程师只能看到自己的报销
        query = Expense.query.filter_by(user_id=current_user.id)
    
    # 应用过滤条件
    if user_id and current_user.has_full_access():
        query = query.filter_by(user_id=user_id)
    
    if expense_type:
        query = query.filter_by(expense_type=expense_type)
    
    if project_id:
        query = query.filter_by(project_id=project_id)
    
    if submit_date_from:
        try:
            from_date = datetime.strptime(submit_date_from, '%Y-%m-%d').date()
            query = query.filter(db.func.date(Expense.submit_date) >= from_date)
        except ValueError:
            pass
    
    if submit_date_to:
        try:
            to_date = datetime.strptime(submit_date_to, '%Y-%m-%d').date()
            query = query.filter(db.func.date(Expense.submit_date) <= to_date)
        except ValueError:
            pass
    
    # 应用费用类别过滤条件
    if expense_category:
        # 通过ExpenseItem的category字段过滤
        expense_ids_with_category = db.session.query(ExpenseItem.expense_id).filter(
            ExpenseItem.category == expense_category
        ).distinct().all()
        expense_ids = [item[0] for item in expense_ids_with_category]
        query = query.filter(Expense.id.in_(expense_ids)) if expense_ids else query.filter(False)
    
    # 获取筛选后的报销列表
    expenses = query.order_by(Expense.submit_date.desc()).all()
    
    # 计算过滤后的总金额
    total_amount = sum(float(expense.total_amount) for expense in expenses)
    
    # 为每个报销获取其对应的处理中任务（获取当前审核人）
    for expense in expenses:
        pending_task = Task.query.filter(
            Task.expense_id == expense.id,
            Task.task_type == 'expense_process',
            Task.status.in_(['处理中', '等待退款', '退款完成'])
        ).first()
        expense.pending_task = pending_task
    
    # 获取汇总数据
    all_users = User.query.all()
    all_projects = Project.query.all()
    all_expense_types = ['项目费用', '售前费用', '其他费用']
    all_expense_categories = [
        '业务招待费', '办公用品', '差旅费', '通讯费', '外包费（对私）', 
        '福利费', '电脑网络打印机硬件费', '培训费', '商务费', '团建费', '财务和审计', '其他费用'
    ]
    
    return render_template(
        "expense_list.html", 
        expenses=expenses,
        all_users=all_users,
        all_projects=all_projects,
        all_expense_types=all_expense_types,
        all_expense_categories=all_expense_categories,
        selected_user_id=user_id,
        selected_expense_type=expense_type,
        selected_project_id=project_id,
        selected_expense_category=expense_category,
        submit_date_from=submit_date_from,
        submit_date_to=submit_date_to,
        total_amount=total_amount
    )

@app.route("/expense/create", methods=["GET", "POST"])
@login_required
def expense_create():
    """创建报销"""
    form = ExpenseForm()
    
    # 设置项目选择列表
    if current_user.has_full_access():
        # 管理员和财务可以看到所有项目
        projects = Project.query.all()
    elif current_user.is_project_manager():
        # 项目经理只能看到关联的项目
        accessible_project_ids = current_user.get_accessible_project_ids()
        projects = Project.query.filter(Project.id.in_(accessible_project_ids)).all() if accessible_project_ids else []
    else:
        # 开发工程师只能看到被分配的项目
        projects = Project.query.join(ProjectAssignment).filter(
            ProjectAssignment.user_id == current_user.id
        ).all()
    
    form.project_id.choices = [('', '请选择项目')] + [(p.id, p.name) for p in projects]
    
    # 如果从项目页面跳转过来，预选择项目
    if request.method == 'GET' and request.args.get('project_id'):
        try:
            preselected_project_id = int(request.args.get('project_id'))
            # 验证用户是否有权限选择这个项目
            if current_user.is_admin() or any(p.id == preselected_project_id for p in projects):
                form.project_id.data = preselected_project_id
        except (ValueError, TypeError):
            pass  # 忽略无效的project_id

    selected_project_id = form.project_id.data if form.project_id.data else None
    form.approver_id.choices = build_approver_choices(selected_project_id)
    
    if form.validate_on_submit():
        if not form.project_id.data:
            flash("请先选择需要关联的项目", "error")
            return render_template("expense_create.html", form=form)
        
        if not form.approver_id.data:
            flash("请选择审批人")
            return render_template("expense_create.html", form=form)

        # 防止重复提交：检查是否在最近10秒内已经创建了相同的报销
        recent_cutoff = datetime.now() - timedelta(seconds=10)
        duplicate_check = Expense.query.filter(
            Expense.user_id == current_user.id,
            Expense.title == form.title.data,
            Expense.total_amount == form.amount.data,
            Expense.created_at >= recent_cutoff
        ).first()
        
        if duplicate_check:
            logger.warning(f"重复提交检测：用户 {current_user.id} 在10秒内尝试提交相同的报销")
            flash("检测到重复提交，请勿重复上传")
            return redirect(url_for("expense_list"))

        # 处理文件上传
        receipt_filename = None
        original_filename = None
        if form.receipt_image.data:
            file = form.receipt_image.data
            if file and not allowed_receipt_file(file.filename):
                flash("报销材料仅支持ZIP格式")
                return render_template("expense_create.html", form=form)
            if file and allowed_receipt_file(file.filename):
                import uuid
                import os
                
                # 保存完整的原始文件名（包括中文）
                original_filename = file.filename
                # 生成唯一文件名，使用原始文件的扩展名
                file_ext = os.path.splitext(file.filename)[1]
                receipt_filename = f"{uuid.uuid4().hex}{file_ext}"
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], receipt_filename)
                file.save(file_path)
        
        # 创建报销单
        expense = Expense(
            user_id=current_user.id,
            project_id=form.project_id.data if form.project_id.data else None,
            title=form.title.data,
            expense_type=form.expense_type.data,
            total_amount=form.amount.data,
            description=form.description.data,
            status='新建'  # 新建时状态为新建
        )
        db.session.add(expense)
        db.session.flush()  # 获取expense.id
        
        # 创建费用明细
        expense_item = ExpenseItem(
            expense_id=expense.id,
            item_name=form.title.data,  # 使用报销标题作为费用明细名称
            category=form.category.data,
            amount=form.amount.data,
            expense_date=form.expense_date.data,
            start_date=form.start_date.data,
            end_date=form.end_date.data,
            description=form.item_note.data,
            receipt_image=receipt_filename,
            receipt_original_name=original_filename
        )
        db.session.add(expense_item)

        task = Task(
            title=f"审批报销：{expense.title}",
            description=f"报销金额：¥{expense.total_amount}\n报销说明：{expense.description or '无'}",
            task_type='expense_process',
            assigned_to=form.approver_id.data,
            assigned_by=current_user.id,
            expense_id=expense.id,
            priority='普通'
        )
        # 验证任务
        task.validate_expense_process_task()
        db.session.add(task)
        
        db.session.commit()
        flash("报销申请提交成功，等待审批")
        return redirect(url_for("expense_list"))
    
    return render_template("expense_create.html", form=form)

@app.route("/expense/<int:expense_id>/toggle_refund", methods=["POST"])
@login_required
def expense_toggle_refund(expense_id):
    """切换报销状态为'已完成'和'退款完成'"""
    expense = Expense.query.get_or_404(expense_id)
    
    # 权限检查 - 只有管理员和财务可以进行此操作
    if not current_user.has_full_access():
        flash("只有管理员或财务可以进行此操作")
        return redirect(url_for("expense_detail", expense_id=expense_id))
    
    # 只能在'已完成'或'退款完成'状态下进行转换
    if expense.status not in ['已完成', '退款完成']:
        flash("只能在已完成或退款完成状态下进行此操作")
        return redirect(url_for("expense_detail", expense_id=expense_id))
    
    # 切换状态
    if expense.status == '已完成':
        expense.status = '退款完成'
        status_msg = '已标记为退款完成'
    else:
        expense.status = '已完成'
        status_msg = '已标记为已完成'
    
    db.session.commit()
    
    log_operation(
        '状态更新',
        '报销',
        f'将报销 "{expense.title}" 状态更新为 {expense.status}',
        'expense',
        expense.id
    )
    
    flash(f"报销{status_msg}")
    return redirect(url_for("expense_detail", expense_id=expense_id))

@app.route("/expense/<int:expense_id>")
@login_required
def expense_detail(expense_id):
    """报销详情"""
    expense = Expense.query.get_or_404(expense_id)
    
    # 权限检查
    can_view = False
    if current_user.has_full_access():
        # 管理员和财务可以查看所有
        can_view = True
    elif expense.user_id == current_user.id:
        # 用户可以查看自己的报销
        can_view = True
    elif current_user.is_project_manager() and expense.project_id:
        # 项目经理可以查看关联项目的报销
        can_view = current_user.can_view_project(expense.project_id)
    
    if not can_view:
        flash("您没有权限查看此报销")
        return redirect(url_for("expense_list"))
    
    # 检查是否有未完成的报销处理任务（如果有，说明已分配给其他人）
    pending_task = Task.query.filter(
        Task.expense_id == expense.id,
        Task.task_type == 'expense_process',
        Task.status.in_(['处理中', '等待退款', '退款完成'])
    ).first()
    
    # 对费用明细中的文件进行去重处理（防止显示重复的文件）
    if expense.items:
        item = expense.items[0]
        if item.receipt_image:
            files = item.receipt_image.split(';')
            names = item.receipt_original_name.split(';') if item.receipt_original_name else []
            
            # 去除重复的文件（保持顺序，只保留第一次出现的）
            unique_files = []
            unique_names = []
            seen_names = set()
            
            for f, n in zip(files, names):
                n_stripped = n.strip()
                if n_stripped and n_stripped not in seen_names:
                    seen_names.add(n_stripped)
                    unique_files.append(f)
                    unique_names.append(n)
            
            # 如果有重复，更新费用明细（可选，取决于是否要持久化修复）
            if len(unique_files) < len(files):
                # 检测到重复文件，更新数据库以修复这个问题
                item.receipt_image = ";".join(unique_files)
                item.receipt_original_name = ";".join(unique_names)
                db.session.commit()
    
    return render_template("expense_detail.html", expense=expense, pending_task=pending_task)

@app.route("/expense/<int:expense_id>/approve", methods=["GET", "POST"])
@app.route("/expenses/<int:expense_id>/approve", methods=["GET", "POST"])
@login_required
def expense_approve(expense_id):
    """审批报销"""
    expense = Expense.query.get_or_404(expense_id)

    can_approve = current_user.has_full_access()
    if (not can_approve and current_user.is_project_manager()
            and expense.project_id
            and current_user.can_view_project(expense.project_id)):
        can_approve = True

    if not can_approve:
        flash("您没有权限审批报销")
        return redirect(url_for("expense_list"))
    
    if expense.status not in ['新建', '处理中', '等待退款', '退款完成']:
        flash("该报销已经审批过了")
        return redirect(url_for("expense_detail", expense_id=expense_id))
    
    form = ExpenseApprovalForm()
    
    # 设置下一个处理人选择列表（除了当前用户和申请人）
    all_users = User.query.filter(User.id != current_user.id).all()
    form.assign_to.choices = [('', '不分配')] + [(u.id, f"{u.username} ({u.role})") for u in all_users]
    
    if form.validate_on_submit():
        if form.approval_action.data == 'approve':
            # 同意审批
            if form.assign_to.data:
                # 有选择下一个处理人，创建新任务
                # 清除之前的所有待处理任务
                old_tasks = Task.query.filter(
                    Task.expense_id == expense.id,
                    Task.task_type == 'expense_process',
                    Task.status.in_(['处理中', '等待退款', '退款完成'])
                ).all()
                for task in old_tasks:
                    task.status = '已完成'
                    task.completed_at = datetime.now()
                
                assigned_user = db.session.get(User, form.assign_to.data)
                task = Task(
                    title=f"处理报销：{expense.title}",
                    description=f"报销金额：¥{expense.total_amount}\n报销说明：{expense.description or '无'}\n审批意见：{form.approve_comment.data or '无'}",
                    task_type='expense_process',
                    assigned_to=form.assign_to.data,
                    assigned_by=current_user.id,
                    expense_id=expense.id,
                    priority='普通'
                )
                task.validate_expense_process_task()
                db.session.add(task)
                expense.status = '处理中'
                expense.approve_comment = form.approve_comment.data
                expense.approver_id = current_user.id
                expense.approve_date = datetime.now()
                db.session.commit()
                
                log_operation(
                    '审批',
                    '报销',
                    f'同意报销 "{expense.title}"，已分配给 {assigned_user.username} 继续处理',
                    'expense',
                    expense.id
                )
                flash(f"报销已同意，已分配给 {assigned_user.username} 处理")
            else:
                # 没有下一步处理人，流程结束
                expense.status = '已完成'
                expense.approve_comment = form.approve_comment.data
                expense.approver_id = current_user.id
                expense.approve_date = datetime.now()
                
                # 清除关联的任务
                pending_tasks = Task.query.filter(
                    Task.expense_id == expense.id,
                    Task.task_type == 'expense_process',
                    Task.status.in_(['处理中', '等待退款', '退款完成'])
                ).all()
                for task in pending_tasks:
                    task.status = '已完成'
                    task.completed_at = datetime.now()
                
                db.session.commit()
                
                log_operation(
                    '审批',
                    '报销',
                    f'同意报销 "{expense.title}"，流程结束',
                    'expense',
                    expense.id
                )
                flash("报销已同意，审批流程结束")
        elif form.approval_action.data == 'waiting_refund':
            # 标记为等待退款
            expense.status = '等待退款'
            expense.approve_comment = form.approve_comment.data
            expense.approver_id = current_user.id
            expense.approve_date = datetime.now()
            
            # 更新相关的任务状态为等待退款
            pending_tasks = Task.query.filter(
                Task.expense_id == expense.id,
                Task.task_type == 'expense_process',
                Task.status == '处理中'
            ).all()
            for task in pending_tasks:
                task.status = '等待退款'
            
            db.session.commit()
            
            log_operation(
                '审批',
                '报销',
                f'报销 "{expense.title}" 标记为等待退款',
                'expense',
                expense.id
            )
            flash(f"报销已标记为等待退款")
        
        elif form.approval_action.data == 'refund_complete':
            # 标记为退款完成（流程结束）
            expense.status = '退款完成'
            expense.approve_comment = form.approve_comment.data
            expense.approver_id = current_user.id
            expense.approve_date = datetime.now()
            
            # 清除关联的任务
            pending_tasks = Task.query.filter(
                Task.expense_id == expense.id,
                Task.task_type == 'expense_process',
                Task.status.in_(['处理中', '等待退款', '退款完成'])
            ).all()
            for task in pending_tasks:
                task.status = '已完成'
                task.completed_at = datetime.now()
            
            db.session.commit()
            
            log_operation(
                '审批',
                '报销',
                f'报销 "{expense.title}" 标记为退款完成',
                'expense',
                expense.id
            )
            flash(f"报销已标记为退款完成")
        else:
            # 拒绝审批，指定申请人为下一步处理人
            applicant = db.session.get(User, expense.user_id)
            task = Task(
                title=f"处理报销：{expense.title}",
                description=f"您的报销申请被拒绝。报销金额：¥{expense.total_amount}\n拒绝原因：{form.approve_comment.data or '无'}\n请根据意见重新处理",
                task_type='expense_process',
                assigned_to=expense.user_id,
                assigned_by=current_user.id,
                expense_id=expense.id,
                priority='普通'
            )
            task.validate_expense_process_task()
            db.session.add(task)
            expense.status = '新建'
            expense.approve_comment = f"被 {current_user.username} 拒绝，原因：{form.approve_comment.data or '无'}"
            expense.approver_id = current_user.id
            expense.approve_date = datetime.now()
            db.session.commit()
            
            log_operation(
                '拒绝',
                '报销',
                f'拒绝报销 "{expense.title}"，已分配给申请人 {applicant.username} 重新处理',
                'expense',
                expense.id
            )
            flash(f"报销已拒绝，已指派给申请人 {applicant.username} 重新处理")
        
        return redirect(url_for("expense_detail", expense_id=expense_id))
    
    return render_template("expense_approve.html", expense=expense, form=form)

@app.route("/expense/<int:expense_id>/edit", methods=["GET", "POST"])
@login_required
def expense_edit(expense_id):
    """编辑报销（新建状态可以编辑）"""
    expense = Expense.query.get_or_404(expense_id)
    
    # 权限检查
    if expense.user_id != current_user.id:
        flash("您只能编辑自己的报销")
        return redirect(url_for("expense_list"))
    
    if expense.status not in ['新建']:
        flash("只有新建的报销可以编辑")
        return redirect(url_for("expense_detail", expense_id=expense_id))
    
    form = ExpenseForm(obj=expense)
    
    # 设置项目选择列表
    if current_user.has_full_access():
        # 管理员和财务可以看到所有项目
        projects = Project.query.all()
    elif current_user.is_project_manager():
        # 项目经理只能看到关联的项目
        accessible_project_ids = current_user.get_accessible_project_ids()
        projects = Project.query.filter(Project.id.in_(accessible_project_ids)).all() if accessible_project_ids else []
    else:
        # 开发工程师只能看到被分配的项目
        projects = Project.query.join(ProjectAssignment).filter(
            ProjectAssignment.user_id == current_user.id
        ).all()
    
    form.project_id.choices = [('', '请选择项目')] + [(p.id, p.name) for p in projects]
    form.approver_id.choices = build_approver_choices(expense.project_id)
    
    # 设置表单的默认值
    if not form.is_submitted():
        form.project_id.data = expense.project_id if expense.project_id else ''
        form.approver_id.data = ''
    
    # 填充表单数据
    if expense.items:
        first_item = expense.items[0]
        if not form.is_submitted():
            form.category.data = first_item.category
            form.amount.data = first_item.amount
            form.expense_date.data = first_item.expense_date
            form.start_date.data = first_item.start_date
            form.end_date.data = first_item.end_date
            form.item_note.data = first_item.description
    
    if form.validate_on_submit():
        if not form.project_id.data:
            flash("请先选择需要关联的项目", "error")
            return render_template("expense_edit.html", form=form, expense=expense)
        
        # 如果原来状态是"已拒绝"，需要选择下一个审批人
        if expense.status == '已拒绝' and not form.approver_id.data:
            flash("重新提交报销必须选择审批人")
            return render_template("expense_edit.html", form=form, expense=expense)
        
        # 处理文件上传
        receipt_filename = None
        original_filename = None
        old_receipt_filenames = None
        old_receipt_names = None
        
        # 保存旧文件名用于追加
        if expense.items and expense.items[0].receipt_image:
            old_receipt_filenames = expense.items[0].receipt_image
            old_receipt_names = expense.items[0].receipt_original_name
        
        if form.receipt_image.data:
            file = form.receipt_image.data
            if file and not allowed_receipt_file(file.filename):
                flash("报销材料仅支持ZIP格式")
                return render_template("expense_edit.html", form=form, expense=expense)
            if file and allowed_receipt_file(file.filename):
                import uuid
                import os
                
                # 保存完整的原始文件名（包括中文）
                original_filename = file.filename
                # 生成唯一文件名，使用原始文件的扩展名
                file_ext = os.path.splitext(file.filename)[1]
                receipt_filename = f"{uuid.uuid4().hex}{file_ext}"
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], receipt_filename)
                file.save(file_path)
                
                # 不删除旧文件，而是追加到列表中（检查重复，避免追加相同的文件）
                new_filename_stripped = file.filename.strip()
                should_append = True
                
                if old_receipt_names:
                    # 检查新文件名是否已经存在
                    existing_names = [n.strip() for n in old_receipt_names.split(';')]
                    if new_filename_stripped in existing_names:
                        should_append = False
                        flash(f"文件 {file.filename} 已经存在，不重复添加")
                
                if should_append:
                    if old_receipt_filenames:
                        receipt_filename = old_receipt_filenames + ";" + receipt_filename
                    if old_receipt_names:
                        original_filename = old_receipt_names + ";" + original_filename
                else:
                    # 重复文件不追加，重置文件名变量，保留原有文件列表
                    receipt_filename = None
                    original_filename = None
        
        # 更新报销单
        expense.title = form.title.data
        expense.expense_type = form.expense_type.data
        expense.project_id = form.project_id.data if form.project_id.data else None
        expense.total_amount = form.amount.data
        expense.description = form.description.data
        
        # 如果原来状态是"已拒绝"，编辑后改为"待审批"并重新分配任务
        was_rejected = expense.status == '已拒绝'
        if was_rejected:
            expense.status = '待审批'
            expense.approver_id = None
            expense.approve_date = None
            # 保留拒绝原因，以便申请人了解为何被拒绝
            # expense.approve_comment = None  # 不清除，保留历史记录
            
            # 取消旧的任务，创建新的待处理任务
            old_tasks = Task.query.filter(
                Task.expense_id == expense.id,
                Task.task_type == 'expense_process',
                Task.status.notin_(['已完成', '拒绝'])
            ).all()
            for task in old_tasks:
                task.status = '拒绝'
                task.completed_at = datetime.now()
            
            # 创建新的待处理任务
            if form.approver_id.data:
                create_expense_task(
                    title=f"审批报销：{expense.title}",
                    description=f"报销金额：¥{expense.total_amount}\n报销说明：{expense.description or '无'}",
                    assigned_to=form.approver_id.data,
                    assigned_by=current_user.id,
                    expense_id=expense.id,
                    priority='普通'
                )
        
        # 更新费用明细（简化处理，只更新第一条）
        if expense.items:
            first_item = expense.items[0]
            first_item.item_name = form.title.data  # 更新费用明细名称
            first_item.category = form.category.data
            first_item.amount = form.amount.data
            first_item.expense_date = form.expense_date.data
            first_item.start_date = form.start_date.data
            first_item.end_date = form.end_date.data
            first_item.description = form.item_note.data
            
            # 如果上传了新文件，更新文件名和原始文件名
            if receipt_filename:
                first_item.receipt_image = receipt_filename
                first_item.receipt_original_name = original_filename
        
        db.session.commit()
        flash("报销修改成功")
        return redirect(url_for("expense_detail", expense_id=expense_id))
    
    return render_template("expense_edit.html", form=form, expense=expense)

@app.route("/expense/<int:expense_id>/delete", methods=["POST"])
@login_required
def expense_delete(expense_id):
    """删除报销（新建状态可以删除）"""
    expense = Expense.query.get_or_404(expense_id)
    
    # 权限检查 - 只有申请人或管理员可以删除
    if expense.user_id != current_user.id and not current_user.has_full_access():
        flash("您没有权限删除此报销")
        return redirect(url_for("expense_list"))
    
    # 只有"新建"状态的报销可以删除
    if expense.status != '新建':
        flash("只有新建状态的报销可以删除")
        return redirect(url_for("expense_detail", expense_id=expense_id))
    
    # 删除关联的任务
    related_tasks = Task.query.filter_by(expense_id=expense_id).all()
    for task in related_tasks:
        db.session.delete(task)
    
    # 删除报销及其关联的文件
    for item in expense.items:
        if item.receipt_image:
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], item.receipt_image)
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
            except Exception as e:
                logger.error(f"删除文件失败: {str(e)}")
    
    db.session.delete(expense)
    db.session.commit()
    
    log_operation(
        '删除',
        '报销',
        f'删除报销 "{expense.title}"，金额：¥{expense.total_amount}，同时删除了 {len(related_tasks)} 个关联任务',
        'expense',
        expense_id
    )
    
    flash("报销已删除")
    return redirect(url_for("expense_list"))

@app.route("/expense/download/<int:item_id>")
@login_required
def expense_download(item_id):
    """下载报销材料，使用原始文件名"""
    item = ExpenseItem.query.get_or_404(item_id)
    file_index = request.args.get('file_index', 0, type=int)
    
    # 权限检查 - 允许：
    # 1. 报销单的提交者
    # 2. 管理员和财务
    # 3. 项目经理（如果报销单关联项目且项目经理是该项目的管理者）
    expense = item.expense
    can_download = False
    
    if expense.user_id == current_user.id:
        can_download = True
    elif current_user.has_full_access():
        can_download = True
    elif current_user.is_project_manager() and expense.project_id and current_user.can_view_project(expense.project_id):
        can_download = True
    
    if not can_download:
        flash("您没有权限下载此文件")
        return redirect(url_for("expense_list"))
    
    if not item.receipt_image:
        flash("该报销明细没有上传文件")
        return redirect(url_for("expense_detail", expense_id=expense.id))
    
    # 支持多文件，用;分隔
    receipt_files = item.receipt_image.split(';')
    receipt_names = item.receipt_original_name.split(';') if item.receipt_original_name else []
    
    # 检查索引是否有效
    if file_index >= len(receipt_files):
        flash("文件不存在")
        return redirect(url_for("expense_detail", expense_id=expense.id))
    
    receipt_file = receipt_files[file_index].strip()
    if not receipt_file:
        flash("文件不存在")
        return redirect(url_for("expense_detail", expense_id=expense.id))
    
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], receipt_file)
    
    if not os.path.exists(file_path):
        flash("文件不存在")
        return redirect(url_for("expense_detail", expense_id=expense.id))
    
    # 使用原始文件名，如果没有则使用UUID文件名
    download_name = receipt_names[file_index].strip() if file_index < len(receipt_names) else receipt_file
    
    return send_file(
        file_path,
        as_attachment=True,
        download_name=download_name
    )

@app.route("/expense/view/<int:item_id>")
@login_required
def expense_view(item_id):
    """在浏览器中查看报销材料，使用原始文件名"""
    item = ExpenseItem.query.get_or_404(item_id)
    file_index = request.args.get('file_index', 0, type=int)
    
    # 权限检查 - 允许：
    # 1. 报销单的提交者
    # 2. 管理员和财务
    # 3. 项目经理（如果报销单关联项目且项目经理是该项目的管理者）
    expense = item.expense
    can_view = False
    
    if expense.user_id == current_user.id:
        can_view = True
    elif current_user.has_full_access():
        can_view = True
    elif current_user.is_project_manager() and expense.project_id and current_user.can_view_project(expense.project_id):
        can_view = True
    
    if not can_view:
        flash("您没有权限查看此文件")
        return redirect(url_for("expense_list"))
    
    if not item.receipt_image:
        flash("该报销明细没有上传文件")
        return redirect(url_for("expense_detail", expense_id=expense.id))
    
    # 支持多文件，用;分隔
    receipt_files = item.receipt_image.split(';')
    receipt_names = item.receipt_original_name.split(';') if item.receipt_original_name else []
    
    # 检查索引是否有效
    if file_index >= len(receipt_files):
        flash("文件不存在")
        return redirect(url_for("expense_detail", expense_id=expense.id))
    
    receipt_file = receipt_files[file_index].strip()
    if not receipt_file:
        flash("文件不存在")
        return redirect(url_for("expense_detail", expense_id=expense.id))
    
    file_ext = os.path.splitext(receipt_file)[1].lower().lstrip('.')
    if file_ext != 'zip':
        flash("查看报销材料仅支持ZIP格式")
        return redirect(url_for("expense_detail", expense_id=expense.id))

    file_path = os.path.join(app.config['UPLOAD_FOLDER'], receipt_file)
    
    if not os.path.exists(file_path):
        flash("文件不存在")
        return redirect(url_for("expense_detail", expense_id=expense.id))
    
    # 使用原始文件名，如果没有则使用UUID文件名
    download_name = receipt_names[file_index].strip() if file_index < len(receipt_names) else receipt_file
    
    return send_file(
        file_path,
        as_attachment=False,  # 不作为附件，直接在浏览器中显示
        download_name=download_name
    )

@app.route("/expense/receipt/delete/<int:item_id>/<int:file_index>", methods=["POST"])
@login_required
def delete_receipt_file(item_id, file_index):
    """删除报销明细中的指定文件"""
    item = ExpenseItem.query.get_or_404(item_id)
    expense = item.expense
    
    # 权限检查 - 只有报销单的提交者才能删除文件
    if expense.user_id != current_user.id:
        return jsonify({"success": False, "message": "您没有权限删除此文件"})
    
    # 检查报销单状态 - 只有"新建"状态才能删除文件
    if expense.status != '新建':
        return jsonify({"success": False, "message": "只有新建的报销才能删除文件"})
    
    if not item.receipt_image:
        return jsonify({"success": False, "message": "该报销明细没有文件"})
    
    # 支持多文件，用;分隔
    receipt_files = item.receipt_image.split(';')
    receipt_names = item.receipt_original_name.split(';') if item.receipt_original_name else []
    
    # 检查索引是否有效
    if file_index >= len(receipt_files):
        return jsonify({"success": False, "message": "文件不存在"})
    
    receipt_file = receipt_files[file_index].strip()
    
    # 删除物理文件
    if receipt_file:
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], receipt_file)
        if os.path.exists(file_path):
            try:
                os.remove(file_path)
            except Exception as e:
                logger.error(f"删除文件失败: {e}")
                return jsonify({"success": False, "message": f"删除文件失败: {str(e)}"})
    
    # 从列表中删除该文件
    receipt_files.pop(file_index)
    if file_index < len(receipt_names):
        receipt_names.pop(file_index)
    
    # 更新数据库
    item.receipt_image = ";".join(receipt_files) if receipt_files else None
    item.receipt_original_name = ";".join(receipt_names) if receipt_names else None
    
    db.session.commit()
    
    return jsonify({"success": True, "message": "文件删除成功"})

# 任务管理路由
@app.route("/tasks")
@login_required
def task_list():
    """我的任务"""
    if current_user.is_admin():
        # 管理员可以看到所有未完成的任务
        # 过滤：排除没有有效报销关联的expense_process类型任务
        tasks = Task.query.filter(
            Task.status.in_(['处理中', '等待退款'])
        ).order_by(Task.created_at.desc()).all()
        # 过滤孤立任务（expense_process类型但expense_id为空或关联报销已删除）
        tasks = [t for t in tasks if not (t.task_type == 'expense_process' and (not t.expense_id or not db.session.get(Expense, t.expense_id)))]
    elif current_user.is_finance():
        # 财务可以看到所有未完成的任务
        tasks = Task.query.filter(
            Task.status.in_(['处理中', '等待退款'])
        ).order_by(Task.created_at.desc()).all()
        # 过滤孤立任务
        tasks = [t for t in tasks if not (t.task_type == 'expense_process' and (not t.expense_id or not db.session.get(Expense, t.expense_id)))]
    elif current_user.is_project_manager():
        # 项目经理可以看到分配给自己的未完成任务
        assigned_tasks = Task.query.filter(
            Task.assigned_to == current_user.id,
            Task.status.in_(['处理中', '等待退款'])
        ).all()
        assigned_ids = {t.id for t in assigned_tasks}
        
        accessible_project_ids = current_user.get_accessible_project_ids()
        expense_tasks = []
        
        if accessible_project_ids:
            # 查询所有expense_process类型的未完成任务
            all_expense_tasks = Task.query.filter(
                Task.task_type == 'expense_process',
                Task.status.in_(['处理中', '等待退款']),
                Task.expense_id.isnot(None)
            ).all()
            
            # 筛选出对应项目的任务
            for task in all_expense_tasks:
                expense = db.session.get(Expense, task.expense_id)
                if expense and expense.project_id in accessible_project_ids:
                    expense_tasks.append(task)
        
        task_ids = assigned_ids | {t.id for t in expense_tasks}
        tasks = Task.query.filter(Task.id.in_(task_ids)).order_by(Task.created_at.desc()).all() if task_ids else []
        # 过滤孤立任务
        tasks = [t for t in tasks if not (t.task_type == 'expense_process' and (not t.expense_id or not db.session.get(Expense, t.expense_id)))]
    else:
        # 普通用户只能看到分配给自己的未完成任务
        tasks = Task.query.filter(
            Task.assigned_to == current_user.id,
            Task.status.in_(['处理中', '等待退款'])
        ).order_by(Task.created_at.desc()).all()
        # 过滤孤立任务
        tasks = [t for t in tasks if not (t.task_type == 'expense_process' and (not t.expense_id or not db.session.get(Expense, t.expense_id)))]
    
    return render_template("task_list.html", tasks=tasks)

@app.route("/task/<int:task_id>")
@login_required
def task_detail(task_id):
    """任务详情"""
    task = Task.query.get_or_404(task_id)
    
    can_view = False
    if current_user.is_admin() or current_user.is_finance():
        can_view = True
    elif task.assigned_to == current_user.id:
        can_view = True
    elif current_user.is_project_manager() and task.task_type == 'expense_process' and task.expense_id:
        expense = db.session.get(Expense, task.expense_id)
        if expense and expense.project_id and current_user.can_view_project(expense.project_id):
            can_view = True
    
    if not can_view:
        flash("您没有权限查看此任务")
        return redirect(url_for("task_list"))
    
    return render_template("task_detail.html", task=task)

@app.route("/task/<int:task_id>/update", methods=["GET", "POST"])
@login_required
def task_update(task_id):
    """更新任务状态"""
    task = Task.query.get_or_404(task_id)
    
    can_update = False
    if current_user.is_admin() or current_user.is_finance():
        can_update = True
    elif task.assigned_to == current_user.id:
        can_update = True
    elif current_user.is_project_manager() and task.task_type == 'expense_process' and task.expense_id:
        expense = db.session.get(Expense, task.expense_id)
        if expense and expense.project_id and current_user.can_view_project(expense.project_id):
            can_update = True
    
    if not can_update:
        flash("您没有权限更新此任务")
        return redirect(url_for("task_list"))
    
    form = TaskUpdateForm()
    
    # 设置下一个处理人的选择列表（除了当前用户）
    all_users = User.query.filter(User.id != current_user.id).all()
    form.assign_to.choices = [('', '不分配')] + [(u.id, f"{u.username} ({u.role})") for u in all_users]
    
    if form.validate_on_submit():
        if form.approval_action.data == 'agree':
            # 同意 - 必须选择下一个处理人，报销状态为"处理中"
            if not form.assign_to.data:
                flash("同意操作必须选择下一个处理人")
                return render_template("task_update.html", form=form, task=task)
            
            next_user = db.session.get(User, form.assign_to.data)
            new_task = Task(
                title=task.title,
                description=form.task_comment.data or task.description,
                task_type=task.task_type,
                assigned_to=form.assign_to.data,
                assigned_by=current_user.id,
                expense_id=task.expense_id,
                priority=task.priority
            )
            if task.task_type == 'expense_process':
                new_task.validate_expense_process_task()
            db.session.add(new_task)
            task.status = '已完成'
            task.completed_at = datetime.now()
            
            # 更新报销状态为"处理中"
            if task.task_type == 'expense_process' and task.expense_id:
                expense = db.session.get(Expense, task.expense_id)
                if expense:
                    expense.status = '处理中'
            
            db.session.commit()
            
            log_operation(
                '完成',
                '任务',
                f'同意任务 "{task.title}"，已转交给 {next_user.username}',
                'task',
                task.id
            )
            flash(f"任务已同意，已转交给 {next_user.username}")
        
        elif form.approval_action.data == 'waiting_refund':
            # 等待退款 - 必须选择下一个处理人来处理退款
            if not form.assign_to.data:
                flash("等待退款操作必须选择下一个处理人来处理退款事宜")
                return render_template("task_update.html", form=form, task=task)
            
            next_user = db.session.get(User, form.assign_to.data)
            new_task = Task(
                title=f"处理退款：{task.title.replace('审批报销：', '').replace('处理报销：', '')}",
                description=form.task_comment.data or f"需要处理退款事宜。{task.description or ''}",
                task_type=task.task_type,
                assigned_to=form.assign_to.data,
                assigned_by=current_user.id,
                expense_id=task.expense_id,
                priority=task.priority
            )
            if task.task_type == 'expense_process':
                new_task.validate_expense_process_task()
            db.session.add(new_task)
            task.status = '已完成'
            task.completed_at = datetime.now()
            
            # 更新报销状态为"等待退款"
            if task.task_type == 'expense_process' and task.expense_id:
                expense = db.session.get(Expense, task.expense_id)
                if expense:
                    expense.status = '等待退款'
            
            db.session.commit()
            
            log_operation(
                '完成',
                '任务',
                f'标记任务 "{task.title}" 为等待退款，已转交给 {next_user.username}',
                'task',
                task.id
            )
            flash(f"已标记为等待退款，已转交给 {next_user.username} 处理退款事宜")
        
        elif form.approval_action.data == 'refund_complete':
            # 退款完成 - 不需要选择处理人，流程结束，报销状态变"退款完成"
            task.status = '已完成'
            task.completed_at = datetime.now()
            
            # 更新报销状态为"退款完成"
            if task.task_type == 'expense_process' and task.expense_id:
                expense = db.session.get(Expense, task.expense_id)
                if expense:
                    expense.status = '退款完成'
                    expense.approve_date = datetime.now()
                    
                    log_operation(
                        '审批',
                        '报销',
                        f'标记报销 "{expense.title}" 为退款完成，流程结束',
                        'expense',
                        expense.id
                    )
            
            db.session.commit()
            flash("任务已完成，报销流程已完成")
                
        elif form.approval_action.data == 'complete':
            # 已完成 - 不需要选择下一处理人，流程结束，报销状态变"已完成"
            task.status = '已完成'
            task.completed_at = datetime.now()
            
            # 更新报销状态为"已完成"
            if task.task_type == 'expense_process' and task.expense_id:
                expense = db.session.get(Expense, task.expense_id)
                if expense:
                    expense.status = '已完成'
                    expense.approve_date = datetime.now()
                    
                    # 自动将费用记录到对应项目中
                    if expense.project_id:
                        for item in expense.items:
                            project_record = ProjectExpenseRecord(
                                project_id=expense.project_id,
                                expense_id=expense.id,
                                category=item.category,
                                amount=item.amount,
                                description=f"报销项目：{expense.title} - {item.category}",
                                recorded_by=current_user.id
                            )
                            db.session.add(project_record)
                    
                    log_operation(
                        '审批',
                        '报销',
                        f'审批完成报销 "{expense.title}"，流程结束',
                        'expense',
                        expense.id
                    )
            
            db.session.commit()
            flash("任务已完成，报销流程已完成")
                
        elif form.approval_action.data == 'reject':
            # 拒绝 - 创建任务分配给申请人，报销状态变"新建"
            if task.task_type == 'expense_process' and task.expense_id:
                expense = db.session.get(Expense, task.expense_id)
                if expense:
                    applicant = db.session.get(User, expense.user_id)
                    reject_task = Task(
                        title=f"重新处理报销：{expense.title}",
                        description=f"您的报销申请被拒绝。拒绝原因：{form.task_comment.data or '无'}\n请根据意见重新处理",
                        task_type='expense_process',
                        assigned_to=expense.user_id,
                        assigned_by=current_user.id,
                        expense_id=expense.id,
                        priority='普通'
                    )
                    reject_task.validate_expense_process_task()
                    db.session.add(reject_task)
                    task.status = '已完成'
                    task.completed_at = datetime.now()
                    expense.status = '新建'
                    db.session.commit()
                    
                    log_operation(
                        '拒绝',
                        '任务',
                        f'拒绝任务 "{task.title}"，已分配给申请人 {applicant.username} 重新处理',
                        'task',
                        task.id
                    )
                    flash(f"任务已拒绝，已指派给申请人 {applicant.username} 重新处理")
                else:
                    flash("无法找到关联的报销单")
            else:
                flash("只能对报销处理任务进行拒绝操作")
                return render_template("task_update.html", form=form, task=task)
        
        return redirect(url_for("task_detail", task_id=task_id))
    
    # 设置当前状态作为默认值
    form.approval_action.data = 'agree'
    
    return render_template("task_update.html", form=form, task=task)

# 添加历史记录查询路由
@app.route("/operation-logs")
@login_required
def operation_logs():
    """操作日志列表"""
    if not current_user.is_admin():
        flash("只有管理员可以查看操作日志")
        return redirect(url_for("index"))
    
    # 获取查询参数
    user_id = request.args.get('user_id', type=int)
    operation_type = request.args.get('operation_type')
    operation_module = request.args.get('operation_module')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    page = request.args.get('page', 1, type=int)
    per_page = 50
    
    # 构建查询
    query = OperationLog.query
    
    if user_id:
        query = query.filter_by(user_id=user_id)
    
    if operation_type:
        query = query.filter_by(operation_type=operation_type)
    
    if operation_module:
        query = query.filter_by(operation_module=operation_module)
    
    if date_from:
        try:
            from_date = datetime.strptime(date_from, '%Y-%m-%d')
            query = query.filter(OperationLog.created_at >= from_date)
        except ValueError:
            pass
    
    if date_to:
        try:
            to_date = datetime.strptime(date_to, '%Y-%m-%d')
            to_date = to_date.replace(hour=23, minute=59, second=59)
            query = query.filter(OperationLog.created_at <= to_date)
        except ValueError:
            pass
    
    # 分页
    pagination = query.order_by(OperationLog.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    logs = pagination.items
    
    # 获取筛选选项
    users = User.query.all()
    operation_types = db.session.query(OperationLog.operation_type).distinct().all()
    operation_types = [t[0] for t in operation_types]
    operation_modules = db.session.query(OperationLog.operation_module).distinct().all()
    operation_modules = [m[0] for m in operation_modules]
    
    return render_template(
        "operation_logs.html",
        logs=logs,
        pagination=pagination,
        users=users,
        operation_types=operation_types,
        operation_modules=operation_modules,
        selected_user_id=user_id,
        selected_operation_type=operation_type,
        selected_operation_module=operation_module,
        date_from=date_from,
        date_to=date_to
    )

if __name__ == "__main__":
    app.run(host='0.0.0.0', port=5005, debug=True)
